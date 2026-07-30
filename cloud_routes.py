import ast
import html
import json
import os
import re
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from urllib.parse import quote

import requests
from flask import Flask, redirect, request, session, url_for, jsonify

from xerox_cloud import get_consumables, normalize_consumables

app = Flask(__name__, template_folder="templates")
DEFAULT_LIMIT = None
INVENTORY_EXCEL_PATTERN = "Printer Supplies-Count-Inventario"

PRINTER_READ_URL = "https://office.services.xerox.com/FMP/Printers/AllPrinter/Printer_Read"
TOKEN = ""
REFERER = "https://office.services.xerox.com/FMP/Printers/AllPrinter"
COOKIE = ""
def _load_xerox_auth():
    env_auth = {
        "token": os.getenv("XEROX_TOKEN", ""),
        "cookie": os.getenv("XEROX_COOKIE", ""),
        "referer": os.getenv("XEROX_REFERER", ""),
    }
    if env_auth["token"] and env_auth["cookie"]:
        if not env_auth["referer"]:
            env_auth["referer"] = "https://office.services.xerox.com/FMP/Printers/AllPrinter"
        return env_auth

    auth_path = Path(__file__).with_name("xerox_auth.json")
    if auth_path.exists():
        try:
            data = json.loads(auth_path.read_text(encoding="utf-8"))
        except Exception:
            data = {}
        file_auth = {
            "token": str(data.get("token") or data.get("requestverificationtoken") or ""),
            "cookie": str(data.get("cookie") or ""),
            "referer": str(data.get("referer") or "https://office.services.xerox.com/FMP/Printers/AllPrinter"),
        }
        if file_auth["token"] and file_auth["cookie"]:
            return file_auth

    script_path = Path(__file__).with_name("test_get_printers.py")
    if not script_path.exists():
        return {}

    text = script_path.read_text(encoding="utf-8")
    match = re.search(r"headers\s*=\s*\{(?P<body>.*?)\n\}", text, re.S)
    if not match:
        return {}

    try:
        headers = ast.literal_eval("{" + match.group("body") + "}")
    except Exception:
        return {}

    return {
        "token": headers.get("requestverificationtoken", ""),
        "cookie": headers.get("cookie", ""),
        "referer": headers.get("referer", "https://office.services.xerox.com/FMP/Printers/AllPrinter"),
    }


XEROX_AUTH = _load_xerox_auth()
TOKEN = XEROX_AUTH.get("token", "")
PRINTER_LIST_REFERER = XEROX_AUTH.get("referer", "https://office.services.xerox.com/FMP/Printers/AllPrinter")
PRINTER_DETAILS_REFERER = "https://office.services.xerox.com/FMP/Printers/PrinterDetails?TabName=Status&AssetEncryptedID=mnmLZZjEqrLkFvG8lrcvLKuhW4BVcw1EJG287i3Kfl3WgwAcd%2F0fCSOrL8C5x7VzGpYHBP8hOlqUa7iAsWI%2FHnrxHU2Q6jhOsbjdxmXpe3SqpG8R3tucvlfzgWNHISFOwJq7q%2Fyfz%2Bq7pe3jET4XWi2oTN0boOX3X%2BMe2ZuNmht6OIAMOa09y49O%2BERiww70tZWVBbykGMSDzRBA2xbSM7dCIZTWuKN6qsO5QMFtzB2F6Ey%2BPiuGo9uSXfa%2BbQ3UAYoOr7zvkEIcpVA01oMwDg%3D%3D&DataSorceID=Grid_Printers&AssetAccountID=00000000-0000-0000-0000-000000000000&IsLogRequired=False"
COOKIE = XEROX_AUTH.get("cookie", "")
ASSET_ACCOUNT_ID = "b3d3b9ce-9bed-e811-9679-0025b52f01ef"



def _normalize_match_text(value):
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower()).strip()


def _load_inventory_serial_overrides():
    path = Path(__file__).with_name("inventory_serial_overrides.json")
    if not path.exists():
        return []
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        print(f"[cloud] Unable to read inventory serial overrides: {exc}")
        return []
    return data.get("row_overrides", []) if isinstance(data, dict) else []


def _inventory_override_serial(ip, business="", location="", address="", rules=None):
    if rules is None:
        rules = _load_inventory_serial_overrides()
    ip = str(ip or "").strip()
    row_text = _normalize_match_text(" ".join([business, location, address]))
    for rule in rules:
        rule_ip = str(rule.get("ip") or "").strip()
        serial = str(rule.get("serial") or "").strip()
        if not rule_ip or rule_ip != ip or not serial:
            continue
        rule_location = _normalize_match_text(rule.get("location", ""))
        if rule_location and rule_location not in row_text:
            continue
        return serial
    return ""

def _legacy_inventory_by_ip(exclude_path=None):
    """Read older inventory workbooks only to recover serial numbers by IP."""
    base_dir = Path(__file__).resolve().parent
    exclude = Path(exclude_path).resolve() if exclude_path else None
    candidates = [
        base_dir / "Printer Supplies-Count-Inventario(2-3-2026).xlsx",
        base_dir / "Printer Supplies-Count-Inventario(29-3-25) OK.xlsx",
        base_dir / "Printer Supplies-Count-Inventario.xlsx",
    ]
    by_ip = {}
    try:
        from openpyxl import load_workbook
    except ImportError:
        return by_ip

    def _normalize_header(value):
        return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())

    def _extract_record(row, headers):
        record = {}
        for idx, header in enumerate(headers):
            value = row[idx] if idx < len(row) else ""
            record[_normalize_header(header)] = "" if value is None else str(value).strip()
        return record

    for candidate in candidates:
        try:
            if exclude and candidate.resolve() == exclude:
                continue
        except OSError:
            pass
        if not candidate.exists():
            continue
        try:
            workbook = load_workbook(candidate, data_only=True, read_only=True)
        except Exception:
            continue
        try:
            for sheet in workbook.worksheets:
                headers = None
                for row in sheet.iter_rows(values_only=True):
                    if headers is None:
                        normalized_cells = [_normalize_header(cell) for cell in row]
                        if any("serial" in cell for cell in normalized_cells) or any(cell in {"negocio", "business"} for cell in normalized_cells):
                            headers = [str(cell or "").strip() for cell in row]
                            continue
                        continue
                    if not any(cell not in (None, "") for cell in row):
                        continue
                    row_values = ["" if cell is None else str(cell).strip() for cell in row]
                    record = _extract_record(row, headers)
                    serial = (record.get("serialnumber", "") or record.get("serial", "") or _infer_serial(row_values)).strip()
                    ip = (record.get("ip", "") or record.get("ipaddress", "") or record.get("printerip", "") or _infer_ip(row_values)).strip()
                    if serial and ip and ip not in by_ip:
                        by_ip[ip] = serial
        finally:
            workbook.close()
    return by_ip

def _serial_key(value):
    return re.sub(r"\s+", "", str(value or "").strip()).upper()


def parse_serials(value):
    """Parse serial numbers pasted as lines, commas, semicolons, tabs, or spaces."""
    serials = []
    seen = set()
    for part in re.split(r"[\s,;]+", str(value or "")):
        serial = part.strip()
        key = _serial_key(serial)
        if not key or key in seen:
            continue
        serials.append(serial)
        seen.add(key)
    return serials


def _load_cloud_inventory():
    """Load the minimal, deployment-safe inventory used when Excel is unavailable."""
    path = Path(__file__).with_name("cloud_inventory.json")
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError) as exc:
        print(f"[cloud] Unable to read cloud inventory {path}: {exc}")
        return {}

    inventory = {}
    for serial, entry in (data if isinstance(data, dict) else {}).items():
        clean_serial = str(serial or "").strip()
        if not clean_serial:
            continue
        groups = entry.get("location_groups", []) if isinstance(entry, dict) else []
        group_names = [str(group).strip() for group in groups if str(group).strip()]
        inventory[clean_serial] = {
            "serial": clean_serial,
            "business": " / ".join(group_names),
            "address": "",
            "ip": "",
            "model": "",
            "printer_name": "",
            "location_groups": group_names,
        }
    return inventory

def load_inventory_excel(excel_path=None):
    """Load inventory serials and their business/address metadata from Excel."""
    if excel_path is None:
        base_dir = Path(__file__).resolve().parent
        candidates = [
            base_dir / "Inventario Printers.xlsx",
            base_dir / "Printer Supplies-Count-Inventario(2-3-2026).xlsx",
            base_dir / "Printer Supplies-Count-Inventario(29-3-25) OK.xlsx",
            base_dir / "Printer Supplies-Count-Inventario.xlsx",
        ]
        for candidate in candidates:
            if candidate.exists():
                path = candidate
                break
        else:
            cloud_inventory = _load_cloud_inventory()
            if cloud_inventory:
                return cloud_inventory
            path = candidates[0]
    else:
        path = Path(excel_path)

    if not path.exists():
        print(f"[cloud] Inventory Excel file not found: {path}")
        return {}

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        print(f"[cloud] Unable to load openpyxl: {exc}")
        return {}

    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        print(f"[cloud] Unable to read inventory Excel file {path}: {exc}")
        return {}

    inventory = {}
    serial_by_ip = _legacy_inventory_by_ip(path)
    serial_override_rules = _load_inventory_serial_overrides()

    if path.name.lower() == "inventario printers.xlsx":
        try:
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    row_values = ["" if cell is None else str(cell).strip() for cell in row]
                    ip = row_values[3] if len(row_values) > 3 else ""
                    if not re.fullmatch(r"(?:\d{1,3}\.){3}\d{1,3}", ip):
                        continue
                    address = row_values[0] if len(row_values) > 0 else ""
                    business = row_values[1] if len(row_values) > 1 else ""
                    location = row_values[2] if len(row_values) > 2 else ""
                    model = row_values[4] if len(row_values) > 4 else ""
                    serial = _inventory_override_serial(ip, business, location, address, serial_override_rules) or serial_by_ip.get(ip, "")
                    if not serial:
                        continue
                    inventory[serial] = {
                        "serial": serial,
                        "business": business,
                        "address": address,
                        "ip": ip,
                        "model": model,
                        "printer_name": location,
                    }
            return inventory
        finally:
            workbook.close()

    def _normalize_header(value):
        return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())

    def _extract_record(row, headers):
        record = {}
        for idx, header in enumerate(headers):
            value = row[idx] if idx < len(row) else ""
            record[_normalize_header(header)] = "" if value is None else str(value).strip()
        return record

    for sheet in workbook.worksheets:
        headers = None
        for row in sheet.iter_rows(values_only=True):
            if headers is None:
                normalized_cells = [_normalize_header(cell) for cell in row]
                if any("serial" in cell for cell in normalized_cells) or any(cell in {"negocio", "business"} for cell in normalized_cells):
                    headers = [str(cell or "").strip() for cell in row]
                    continue
                continue

            if not any(cell not in (None, "") for cell in row):
                continue

            record = _extract_record(row, headers)

            row_values = ["" if cell is None else str(cell).strip() for cell in row]

            ip = (
                record.get("ip", "")
                or record.get("ipaddress", "")
                or record.get("printerip", "")
                or _infer_ip(row_values)
            ).strip()
            serial = (
                record.get("serialnumber", "")
                or record.get("serial", "")
                or _infer_serial(row_values)
                or serial_by_ip.get(ip, "")
            ).strip()
            if not serial:
                continue

            business = (
                record.get("business", "")
                or record.get("negocio", "")
                or record.get("customer", "")
                or record.get("company", "")
            ).strip()
            address = (
                record.get("address", "")
                or record.get("direccion", "")
                or record.get("location", "")
            ).strip()
            model = (
                record.get("model", "")
                or record.get("modelname", "")
                or _infer_model(row_values)
            ).strip()
            printer_name = (
                record.get("printername", "")
                or record.get("printer", "")
                or (row_values[0] if row_values else "")
            ).strip()

            inventory[serial] = {
                "serial": serial,
                "business": business,
                "address": address,
                "ip": ip,
                "model": model,
                "printer_name": printer_name,
            }

    workbook.close()
    return inventory



def _inventory_group_names(entry):
    explicit_groups = entry.get("location_groups", [])
    if explicit_groups:
        return list(dict.fromkeys(str(group).strip() for group in explicit_groups if str(group).strip()))
    text = " ".join(str(entry.get(key, "") or "") for key in ("business", "address", "printer_name"))
    normalized = re.sub(r"[^a-z0-9]+", " ", text.lower())
    rules = [
        ("Gallardo", ("gallardo",)),
        ("Flagler", ("flagler", "flager")),
        ("107th", ("107", "avanawest")),
        ("Dallas", ("dallas",)),
        ("Hialeah", ("hialeah",)),
        ("Doral", ("doral",)),
        ("Finca", ("finca", "11970sw64th", "hr yeline", "hr patricia")),
        ("Icon", ("icon",)),
        ("Unique", ("unique",)),
        ("Dale Solution", ("dale solution", "dale",)),
        ("Main Zoo", ("zoo", "olivia",)),
    ]
    groups = [label for label, tokens in rules if any(token in normalized for token in tokens)]
    return groups or ["Other Locations"]


def _inventory_location_payload(inventory):
    serials = []
    seen = set()
    grouped = {}
    for serial, entry in inventory.items():
        key = _serial_key(serial)
        if not key or key in seen:
            continue
        clean_serial = str(serial).strip()
        serials.append(clean_serial)
        seen.add(key)
        for group_name in _inventory_group_names(entry):
            grouped.setdefault(group_name, []).append(clean_serial)

    preferred_order = ["Gallardo", "Flagler", "107th", "Dallas", "Hialeah", "Doral", "Finca", "Icon", "Unique", "Dale Solution", "Main Zoo", "Other Locations"]
    order = {name: idx for idx, name in enumerate(preferred_order)}
    locations = [
        {"name": name, "count": len(values), "serials": values}
        for name, values in sorted(grouped.items(), key=lambda item: (order.get(item[0], len(order)), item[0].lower()))
    ]
    return {"success": True, "serials": serials, "count": len(serials), "locations": locations}

def _infer_serial(values):
    for value in values:
        v = str(value or "").strip()
        if re.fullmatch(r"[A-Z]{2,4}\d{5,8}", v, re.I):
            return v
    return ""


def _infer_ip(values):
    for value in values:
        v = str(value or "").strip()
        if re.fullmatch(r"(?:\d{1,3}\.){3}\d{1,3}", v):
            return v
    return ""


def _infer_model(values):
    for value in values:
        v = str(value or "").strip()
        if re.search(r"\b(?:B415|B7135|B8155|C415)\b", v, re.I):
            return v
    return ""


def _printer_black_impressions(printer):
    for key in ("PageCountMono", "BlackPrintedImpressions", "BlackImpressions", "BlackPageCount", "PageCount"):
        value = printer.get(key)
        if value in (None, ""):
            continue
        try:
            return int(float(str(value).replace(",", "").strip()))
        except (TypeError, ValueError):
            continue
    return -1

def _normalize_printer_entry(printer, inventory=None):
    model_name = printer.get("ModelName", "")
    if not any(model in model_name for model in ["B415", "B7135", "B8155"]):
        return None

    asset_id = printer.get("AssetID")
    asset_encrypted_id = printer.get("AssetEncryptedId")
    if not asset_id or not asset_encrypted_id:
        return normalize_consumables(
            [],
            {
                "model": printer.get("ModelName", ""),
                "serial": printer.get("SerialNumber", ""),
                "ip": printer.get("PrinterIPAddress", ""),
                "black_impressions": _printer_black_impressions(printer),
            },
        )

    result = get_consumables(
        asset_id=asset_id,
        asset_account_id=ASSET_ACCOUNT_ID,
        asset_encrypted_id=asset_encrypted_id,
        token=TOKEN,
        referer=PRINTER_DETAILS_REFERER,
        cookie=COOKIE,
    )

    normalized = normalize_consumables(
        result.get("consumables", []) if result.get("success") else [],
        {
            "model": printer.get("ModelName", ""),
            "serial": printer.get("SerialNumber", ""),
            "ip": printer.get("PrinterIPAddress", ""),
            "black_impressions": _printer_black_impressions(printer),
        },
    )

    if inventory is not None:
        serial = str(printer.get("SerialNumber", "") or "").strip()
        inventory_by_key = {_serial_key(key): value for key, value in inventory.items()}
        inventory_entry = inventory.get(serial, {}) or inventory_by_key.get(_serial_key(serial), {})
        normalized["business"] = inventory_entry.get("business", "")
        normalized["address"] = inventory_entry.get("address", "")

    return normalized


def _format_count(value):
    try:
        count = int(float(str(value).replace(",", "").strip()))
    except (TypeError, ValueError):
        return "N/A"
    if count < 0:
        return "N/A"
    return f"{count:,}"


def _parse_remaining_pct(value):
    if value is None:
        return None
    match = re.search(r"(\d+)", str(value))
    if not match:
        return None
    return int(match.group(1))


def _model_key(model_name):
    name = str(model_name or "").upper()
    for token in ("B8155", "B7135", "B415"):
        if token in name:
            return token
    return "OTHER"


def _printer_status(printer):
    pct_values = [
        _parse_remaining_pct(printer.get("toner")),
        _parse_remaining_pct(printer.get("drum")),
        _parse_remaining_pct(printer.get("fuser")),
        _parse_remaining_pct(printer.get("transfer_roller")),
        _parse_remaining_pct(printer.get("transfer_belt_cleaner")),
        _parse_remaining_pct(printer.get("second_bias_transfer_roll")),
        _parse_remaining_pct(printer.get("waste_toner")),
    ]
    pct_values = [value for value in pct_values if value is not None]
    if any(value < 20 for value in pct_values):
        return "CRITICAL"
    if any(value < 40 for value in pct_values):
        return "WARN"
    return "OK"


def get_project_printers(limit=None, inventory=None, serials=None):
    """Fetch printers from Xerox cloud, filter by serials/inventory, and normalize consumables."""
    if limit is None:
        limit = None
    else:
        try:
            limit = int(limit)
        except (TypeError, ValueError):
            limit = None

    if limit is not None and limit < 1:
        limit = None
    headers = {
        "accept": "application/json",
        "x-requested-with": "XMLHttpRequest",
        "requestverificationtoken": TOKEN,
        "referer": PRINTER_LIST_REFERER,
        "cookie": COOKIE,
        "user-agent": "Mozilla/5.0",
    }
    payload = {
        "page": 1,
        "pageSize": 200,
        "sort[0][field]": "LocationName",
        "sort[0][dir]": "asc",
    }

    try:
        response = requests.post(PRINTER_READ_URL, headers=headers, data=payload, timeout=30)
        response.raise_for_status()
    except requests.HTTPError as exc:
        status_code = getattr(exc.response, "status_code", None)
        if status_code == 401:
            raise RuntimeError("Xerox Cloud authorization expired. Refresh the Xerox Cloud session credentials.") from exc
        raise RuntimeError(f"Xerox Cloud returned HTTP {status_code or 'error'}.") from exc
    except requests.RequestException as exc:
        print(f"[cloud] Error fetching printer list: {exc}")
        raise RuntimeError("Unable to reach Xerox Cloud. Check network access and try again.") from exc

    try:
        data = response.json()
    except ValueError as exc:
        print("[cloud] Invalid JSON response")
        raise RuntimeError("Xerox Cloud returned an invalid response. Refresh the session and try again.") from exc

    candidates = [
        printer for printer in data.get("Data", [])
        if any(model in printer.get("ModelName", "") for model in ["B415", "B7135", "B8155"])
    ]

    if inventory is None:
        inventory = load_inventory_excel()

    selected_serials = parse_serials("\n".join(serials or []))
    selected_keys = {_serial_key(serial) for serial in selected_serials}

    if selected_keys:
        order = {key: idx for idx, key in enumerate(_serial_key(serial) for serial in selected_serials)}
        candidates = [
            printer for printer in candidates
            if _serial_key(printer.get("SerialNumber", "")) in selected_keys
        ]
        candidates.sort(key=lambda printer: order.get(_serial_key(printer.get("SerialNumber", "")), len(order)))
    elif inventory:
        inventory_keys = {_serial_key(serial) for serial in inventory.keys() if _serial_key(serial)}
        candidates = [
            printer for printer in candidates
            if _serial_key(printer.get("SerialNumber", "")) in inventory_keys
        ]
    else:
        candidates = []

    if limit is not None:
        candidates = candidates[:limit]

    printers = []
    if not candidates:
        return printers

    max_workers = min(24, len(candidates))
    def _safe_normalize(printer):
        try:
            return _normalize_printer_entry(printer, inventory=inventory)
        except Exception as exc:
            print(f"[cloud] Error normalizing {printer.get('SerialNumber', '')}: {exc}")
            return normalize_consumables(
                [],
                {
                    "model": printer.get("ModelName", ""),
                    "serial": printer.get("SerialNumber", ""),
                    "ip": printer.get("PrinterIPAddress", ""),
                    "black_impressions": _printer_black_impressions(printer),
                },
            )

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        for normalized in executor.map(_safe_normalize, candidates):
            if normalized is not None:
                printers.append(normalized)

    return printers


def _format_pct(value):
    try:
        pct = float(value)
    except (TypeError, ValueError):
        return ""
    if pct < 0:
        return ""
    if pct.is_integer():
        return f"{int(pct)}%"
    return f"{pct:.1f}%"


def _best_item(items, category):
    candidates = [item for item in items if item.get("category") == category]
    if not candidates:
        return {}
    return max(candidates, key=lambda item: item.get("percent") if isinstance(item.get("percent"), (int, float)) else -1)


def _item_value(item):
    if not item:
        return ""
    status = item.get("status")
    if status:
        return str(status)
    return _format_pct(item.get("percent"))


def _local_snmp_printer(serial, inventory_entry):
    ip = (inventory_entry or {}).get("ip", "")
    if not ip:
        return None

    import app as app_module

    data = app_module.get_cached(
        ip,
        app_module.DEFAULT_COMMUNITY,
        app_module.TIMEOUT_DEFAULT,
        app_module.TTL_DEFAULT,
    )
    items = data.get("items") or []
    model = data.get("model") or inventory_entry.get("model", "")
    transfer_item = _best_item(items, "transfer_roller")
    transfer_value = _item_value(transfer_item)

    row = {
        "model": model,
        "serial": serial,
        "ip": ip,
        "printer_name": data.get("printer_name") or inventory_entry.get("printer_name", ""),
        "black_impressions": data.get("black_impressions", -1),
        "business": inventory_entry.get("business", ""),
        "address": inventory_entry.get("address", ""),
        "toner": _item_value(_best_item(items, "toner")),
        "drum": _item_value(_best_item(items, "drum")),
        "fuser": _item_value(_best_item(items, "fuser")),
        "transfer_roller": transfer_value if "B8155" not in str(model).upper() else "",
        "transfer_belt_cleaner": _item_value(_best_item(items, "belt_cleaner")),
        "second_bias_transfer_roll": transfer_value if "B8155" in str(model).upper() else "",
        "waste_toner": _item_value(_best_item(items, "waste_toner")),
    }
    return row


def get_project_printers_from_inventory_snmp(serials, inventory=None):
    if inventory is None:
        inventory = load_inventory_excel()
    inventory_by_key = {_serial_key(key): value for key, value in inventory.items()}
    targets = []
    for serial in parse_serials("\n".join(serials or [])):
        entry = inventory_by_key.get(_serial_key(serial))
        if entry and entry.get("ip"):
            targets.append((serial, entry))

    if not targets:
        return []

    def _safe_local(target):
        serial, entry = target
        try:
            return _local_snmp_printer(entry.get("serial", serial), entry)
        except Exception as exc:
            print(f"[cloud fallback] SNMP failed for {serial}: {exc}")
            return None

    max_workers = min(12, len(targets))
    printers = []
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        for printer in executor.map(_safe_local, targets):
            if printer:
                printers.append(printer)
    return printers


@app.route("/cloud")
def cloud():
    if not session.get("user_id"):
        return redirect(url_for("login", next=request.full_path if request.query_string else request.path))

    limit_param = request.args.get("limit", default="")
    if limit_param:
        try:
            limit = int(limit_param)
        except ValueError:
            limit = None
    else:
        limit = None

    if limit is not None and limit < 1:
        limit = None

    serials_text = request.args.get("serials", default="")
    requested_serials = parse_serials(serials_text)

    inventory = load_inventory_excel() if requested_serials else {}
    search_error = ""
    try:
        printers = (
            get_project_printers(limit=limit, inventory=inventory, serials=requested_serials)
            if requested_serials
            else []
        )
    except Exception as exc:
        print(f"[cloud] Search failed: {exc}")
        printers = []
        search_error = str(exc) or "Search failed. Check the Xerox Cloud session and try again."
    found_serial_keys = {_serial_key(printer.get("serial", "")) for printer in printers}
    missing_serials = [] if search_error else [
        serial for serial in requested_serials
        if _serial_key(serial) not in found_serial_keys
    ]

    if missing_serials and inventory:
        inventory_by_key = {_serial_key(key): value for key, value in inventory.items()}
        fallback_serials = [
            serial for serial in missing_serials
            if "Finca" in _inventory_group_names(inventory_by_key.get(_serial_key(serial), {}))
        ]
        fallback_printers = get_project_printers_from_inventory_snmp(fallback_serials, inventory=inventory)
        if fallback_printers:
            printers.extend(fallback_printers)
            found_serial_keys = {_serial_key(printer.get("serial", "")) for printer in printers}
            missing_serials = [
                serial for serial in requested_serials
                if _serial_key(serial) not in found_serial_keys
            ]


    rows_html = []
    inventory_error = None
    if not inventory and requested_serials:
        inventory_error = None

    for printer in printers:
        model_name = str(printer.get("model", "") or "")
        model = html.escape(_model_key(model_name))
        ip = html.escape(printer.get("ip", "") or "")
        serial = html.escape(printer.get("serial", "") or "")
        toner = html.escape(printer.get("toner", "") or "")
        drum = html.escape(printer.get("drum", "") or "")
        transfer_roller = html.escape(printer.get("transfer_roller", "") or "")
        transfer_belt_cleaner = html.escape(printer.get("transfer_belt_cleaner", "") or "")
        second_bias_transfer_roll = html.escape(printer.get("second_bias_transfer_roll", "") or "")
        waste_toner = html.escape(printer.get("waste_toner", "") or "")
        fuser = html.escape(printer.get("fuser", "") or "")
        black_impressions = html.escape(_format_count(printer.get("black_impressions", -1)))
        business = html.escape(printer.get("business", "") or "")
        address = html.escape(printer.get("address", "") or "")
        status = _printer_status(printer)
        model_key = _model_key(printer.get("model", ""))
        toner_pct = _parse_remaining_pct(printer.get("toner"))
        drum_pct = _parse_remaining_pct(printer.get("drum"))
        fuser_pct = _parse_remaining_pct(printer.get("fuser"))
        transfer_pct = _parse_remaining_pct(printer.get("transfer_roller"))
        transfer_belt_cleaner_pct = _parse_remaining_pct(printer.get("transfer_belt_cleaner"))
        second_bias_transfer_roll_pct = _parse_remaining_pct(printer.get("second_bias_transfer_roll"))
        waste_toner_pct = _parse_remaining_pct(printer.get("waste_toner"))

        def _metric_cell(value, pct):
            if value in (None, ""):
                return "<span class=\"pill\">N/A</span>"
            try:
                pct_num = int(str(pct).replace('%', '').strip()) if pct is not None else None
            except Exception:
                pct_num = None
            if pct_num is None:
                return f"<span class=\"pill\">{html.escape(str(value))}</span>"
            cls = "meter ok" if pct_num >= 50 else "meter warn" if pct_num >= 20 else "meter low"
            color = "var(--ok)" if pct_num >= 50 else "var(--warn)" if pct_num >= 20 else "var(--bad)"
            return (
                f"<div class=\"{cls}\"><div class=\"bar\"><span style=\"width:{pct_num}%;background:{color}\"></span></div>"
                f"<span class=\"value\">{pct_num}%</span></div>"
            )

        rows_html.append(
            f"<tr data-model=\"{model_key}\" data-status=\"{status}\" data-toner=\"{toner_pct if toner_pct is not None else ''}\" data-drum=\"{drum_pct if drum_pct is not None else ''}\" data-fuser=\"{fuser_pct if fuser_pct is not None else ''}\" data-transfer=\"{transfer_pct if transfer_pct is not None else ''}\" data-transfer-belt-cleaner=\"{transfer_belt_cleaner_pct if transfer_belt_cleaner_pct is not None else ''}\" data-second-bias-transfer-roll=\"{second_bias_transfer_roll_pct if second_bias_transfer_roll_pct is not None else ''}\" data-waste-toner=\"{waste_toner_pct if waste_toner_pct is not None else ''}\" data-business=\"{business}\" data-address=\"{address}\">"
            f"<td><span class=\"pill\">{model}</span></td>"
            f"<td>{serial or 'N/A'}</td>"
            f"<td>{ip}</td>"
            f"<td>{black_impressions}</td>"
            f"<td>{_metric_cell(toner, toner_pct)}</td>"
            f"<td>{_metric_cell(drum, drum_pct)}</td>"
            f"<td>{_metric_cell(fuser, fuser_pct)}</td>"
            f"<td>{_metric_cell(transfer_roller, transfer_pct)}</td>"
            f"<td>{_metric_cell(transfer_belt_cleaner, transfer_belt_cleaner_pct)}</td>"
            f"<td>{_metric_cell(second_bias_transfer_roll, second_bias_transfer_roll_pct)}</td>"
            f"<td>{_metric_cell(waste_toner, waste_toner_pct)}</td>"
            f"<td><span class=\"status {status.lower() if status.lower() in ('ok','warn','critical') else 'ok'}\">{status}</span></td>"
            "</tr>"
        )

    if rows_html:
        rows_html_content = "".join(rows_html)
    else:
        if inventory_error:
            rows_html_content = f"<tr><td colspan='12' class='empty'>{html.escape(inventory_error)}</td></tr>"
        elif search_error:
            rows_html_content = f"<tr><td colspan='12' class='empty'>{html.escape(search_error)}</td></tr>"
        elif requested_serials:
            rows_html_content = "<tr><td colspan='12' class='empty'>No printers found for the entered serial numbers.</td></tr>"
        else:
            rows_html_content = "<tr><td colspan='12' class='empty'>Enter serial numbers and press Search.</td></tr>"

    template_path = Path(__file__).resolve().parent / "templates" / "cloud.html"
    page = template_path.read_text(encoding="utf-8")

    limit_text = f" (limit: {limit})" if limit is not None else ""
    missing_notice = ""
    if search_error:
        missing_notice = '<div class="notice error">' + html.escape(search_error) + "</div>"
    elif missing_serials:
        missing_notice = (
            '<div class="notice">Not found: '
            + html.escape(", ".join(missing_serials))
            + "</div>"
        )
    # Do not auto-populate the serials textarea on page load. Leave population to the user action (button).
    inventory_payload = {"success": True, "serials": [], "count": 0, "locations": []}
    try:
        inventory_payload = _inventory_location_payload(load_inventory_excel())
    except Exception as exc:
        print(f"[cloud] Unable to build inventory location list: {exc}")
    bootstrap = "window.__INVENTORY_DATA__ = " + json.dumps(inventory_payload) + ";"
    username = session.get("username") or "administrator"
    manage_users_link = """        <a href="/users">Manage Users</a>
        <span class="divider">|</span>
"""
    if (username or "").strip().lower() != "joan":
        page = page.replace(manage_users_link, "")
    lang = session.get("lang") if session.get("lang") in {"en", "es"} else "en"
    next_url = request.full_path.rstrip("?") or "/cloud"
    next_arg = quote(next_url, safe="/?=%")

    if lang == "es":
        translations = {
            '<html lang="en">': '<html lang="es">',
            "Printers Supplies": "Suministros de Impresoras",
            "Supply monitoring and fleet health overview": "Monitoreo de suministros y estado de la flota",
            "Signed in as": "Conectado como",
            "Edit Account": "Editar Cuenta",
            "Logout": "Salir",
            "Lookup": "Buscar",
            "Summary": "Resumen",
            "Results": "Resultados",
            "Showing": "Mostrando",
            "devices": "equipos",
            "Clear": "Limpiar",
            "Supply Lookup": "Busqueda de Suministros",
            "Ready": "Listo",
            "Serial numbers": "Numeros de serie",
            "Query": "Buscar",
            "Searching devices...": "Buscando equipos...",
            "Contacting Xerox Cloud": "Contactando Xerox Cloud",
            "Ready to query serial numbers": "Listo para buscar numeros de serie",
            "Total printers": "Total de impresoras",
            "Total in current query": "Total en la busqueda actual",
            "Low supplies": "Suministros bajos",
            "Supplies requiring attention": "Suministros que requieren atencion",
            "Model types": "Tipos de modelo",
            "Unique printer models detected": "Modelos unicos detectados",
            "Last update": "Ultima actualizacion",
            "Local browser time": "Hora local del navegador",
            "Show only below": "Mostrar solo debajo",
            "Search": "Buscar",
            "Filter by serial, IP, name, or model": "Filtrar por serial, IP, nombre o modelo",
            "Model": "Modelo",
            "Serial": "Serial",
            "Toner": "Toner",
            "Drum": "Drum",
            "Fuser": "Fuser",
            "Transfer": "Transfer",
            "Transfer Belt Cleaner": "Limpiador de banda de transferencia",
            "Second Bias Transfer Roll": "Segundo rodillo de transferencia bias",
            "Waste Toner Container": "Contenedor de toner residual",
            "Status": "Estado",
            "Enter serial numbers and press Search to load supplies.": "Ingrese numeros de serie y presione Buscar para cargar suministros.",
            "No printers found for the entered serial numbers.": "No se encontraron impresoras para los numeros de serie ingresados.",
            "Enter serial numbers and press Search.": "Ingrese numeros de serie y presione Buscar.",
            "Not found:": "No encontrados:",
            "Querying...": "Buscando...",
        }
        for source, target in sorted(translations.items(), key=lambda item: len(item[0]), reverse=True):
            page = page.replace(source, target)

    page = (
        page
        .replace('/language/en?next=/cloud', f'/language/en?next={next_arg}')
        .replace('/language/es?next=/cloud', f'/language/es?next={next_arg}')
    )

    return (
        page
        .replace("<!-- CLOUD_USERNAME -->administrator<!-- /CLOUD_USERNAME -->", html.escape(username))
        .replace("<!-- CLOUD_LANG_EN -->", "active" if lang == "en" else "")
        .replace("<!-- CLOUD_LANG_ES -->", "active" if lang == "es" else "")
        .replace('<strong id="countValue">0</strong>', f'<strong id="countValue">{len(printers)}</strong>')
        .replace("<!-- CLOUD_LIMIT -->", html.escape(limit_text))
        .replace("<!-- CLOUD_MISSING_NOTICE -->", missing_notice)
        .replace('<tr><td colspan="12" class="empty">Enter serial numbers and press Search to load supplies.</td></tr>', rows_html_content)
        .replace("<!-- CLOUD_BOOTSTRAP -->", bootstrap)
    )


if __name__ == "__main__":
    app.run(debug=True)

