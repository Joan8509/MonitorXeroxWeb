import io
from datetime import datetime
from typing import Any, Dict, List
from flask import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule

def _styles():
    thin = Side(border_style="thin", color="D3D3D3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "header_font": Font(bold=True, color="FFFFFF"),
        "header_fill": PatternFill("solid", fgColor="2563EB"),
        "center": Alignment(horizontal="center", vertical="center"),
        "border": border
    }

def _auto_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value or "")))
            except:
                pass
        ws.column_dimensions[col_letter].width = max(12, min(max_length + 2, 40))

# ------------------ EXPORT SINGLE PRINTER ------------------
def handle_export_xlsx_pivot(data: Dict[str, Any]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Supplies"
    st = _styles()

    headers = ["IP", "Printer", "Model", "Item", "Percent"]
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col)
        c.font = st["header_font"]
        c.fill = st["header_fill"]
        c.alignment = st["center"]
        c.border = st["border"]

    ip = data.get("ip", "")
    model = data.get("model", "")
    printer = data.get("printer_name", "")
    items = data.get("items", [])

    for it in items:
        ws.append([ip, printer, model, it.get("category", ""), it.get("percent", "")])

    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = st["center"]
            c.border = st["border"]

    rule = DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100,
                       color="4F81BD", showValue="None")
    ws.conditional_formatting.add("E2:E1000", rule)

    _auto_width(ws)
    filename = f"Supplies_{ip}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return Response(
        stream.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ------------------ EXPORT MULTIPLE PRINTERS ------------------
def handle_export_xlsx_list_pivot(data: List[Dict[str, Any]]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Supplies List"
    st = _styles()

    headers = ["IP", "Printer", "Model", "Item", "Percent"]
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col)
        c.font = st["header_font"]
        c.fill = st["header_fill"]
        c.alignment = st["center"]
        c.border = st["border"]

    for d in data:
        ip = d.get("ip", "")
        model = d.get("model", "")
        printer = d.get("printer_name", "")
        for it in d.get("items", []):
            ws.append([ip, printer, model, it.get("category", ""), it.get("percent", "")])

    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = st["center"]
            c.border = st["border"]

    rule = DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100,
                       color="4F81BD", showValue="None")
    ws.conditional_formatting.add("E2:E10000", rule)

    _auto_width(ws)
    filename = f"Supplies_List_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return Response(
        stream.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
