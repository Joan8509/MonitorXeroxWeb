
# app.py
import os, time, io, re, sqlite3, html as htmlmod
from functools import wraps
from pathlib import Path
from typing import Any, Dict, List, Tuple, Optional
from flask import Flask, jsonify, request, Response, session, redirect, url_for
from werkzeug.security import generate_password_hash, check_password_hash
from puresnmp import walk, get as snmp_get

from cloud_routes import cloud as cloud_view, load_inventory_excel, _serial_key

# ---- XLSX pretty export ----
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule, CellIsRule

app = Flask(__name__)
BASE_DIR = Path(__file__).resolve().parent


def _inventory_group_names(entry):
    text = " ".join(str(entry.get(key, "") or "") for key in ("business", "address", "printer_name"))
    normalized = re.sub(r"[^a-z0-9]+", " ", text.lower())
    rules = [
        ("Gallardo", ("gallardo",)),
        ("Flagler", ("flagler", "flager")),
        ("107th", ("107", "avanawest")),
        ("Dallas", ("dallas",)),
        ("Hialeah", ("hialeah",)),
        ("Doral", ("doral",)),
        ("Finca", ("finca", "11970sw64th", "hr yeline", "hr patricia")),
        ("Icon", ("icon",)),
        ("Unique", ("unique",)),
        ("Dale Solution", ("dale solution", "dale",)),
        ("Main Zoo", ("zoo", "olivia",)),
    ]
    groups = [label for label, tokens in rules if any(token in normalized for token in tokens)]
    return groups or ["Other Locations"]


def _inventory_location_payload(inventory):
    serials = []
    seen = set()
    grouped = {}
    for serial, entry in inventory.items():
        key = _serial_key(serial)
        if not key or key in seen:
            continue
        clean_serial = str(serial).strip()
        serials.append(clean_serial)
        seen.add(key)
        for group_name in _inventory_group_names(entry):
            grouped.setdefault(group_name, []).append(clean_serial)

    preferred_order = ["Gallardo", "Flagler", "107th", "Dallas", "Hialeah", "Doral", "Finca", "Icon", "Unique", "Dale Solution", "Main Zoo", "Other Locations"]
    order = {name: idx for idx, name in enumerate(preferred_order)}
    locations = [
        {"name": name, "count": len(values), "serials": values}
        for name, values in sorted(grouped.items(), key=lambda item: (order.get(item[0], len(order)), item[0].lower()))
    ]
    return {"success": True, "serials": serials, "count": len(serials), "locations": locations}

@app.route("/cloud/inventory_serials")
def cloud_inventory_serials():
  # Record every request for diagnostics (authenticated or not)
  try:
    from datetime import datetime
    cookie = request.headers.get('Cookie') or ''
    referer = request.headers.get('Referer') or ''
    now = datetime.utcnow()
    common_line = f"{now.isoformat()}Z\tREMOTE={request.remote_addr}\tCOOKIE_PRESENT={bool(cookie)}\tSESSION_USER={session.get('user_id')}\tSESSION_KEYS={list(session.keys())}\tREFERER={referer}\n"
    try:
      with open('inventory_fetch.log','a',encoding='utf-8') as fh:
        fh.write(common_line)
    except Exception:
      print('[inventory] Unable to write inventory_fetch.log')
  except Exception:
    pass

  if not session.get("user_id"):
    # Also print a clear console message for unauthenticated callers
    cookie = request.headers.get('Cookie') or ''
    print('[inventory] auth required for', request.remote_addr, 'cookie_present=', bool(cookie))
    return jsonify(success=False, error="auth_required", message="Authentication required", cookie_present=bool(cookie)), 401

  inventory = load_inventory_excel()
  return jsonify(_inventory_location_payload(inventory))


@app.route("/_debug_inventory_serials")
def _debug_inventory_serials():
  """Debug endpoint (no auth) to quickly test server routing returns inventory serials."""
  try:
    inventory = load_inventory_excel()
  except Exception:
    inventory = {}
  serials = []
  seen = set()
  for serial in inventory.keys():
    key = _serial_key(serial)
    if not key or key in seen:
      continue
    serials.append(str(serial).strip())
    seen.add(key)
  return jsonify(success=True, serials=serials, count=len(serials))


@app.route("/cloud")
def cloud():
    if not session.get("user_id"):
        return redirect(url_for("login", next=request.full_path if request.query_string else request.path))
    return cloud_view()

@app.route("/p")
def short_cloud():
    return redirect("/cloud")

@app.route("/monitor")
def short_monitor():
    return redirect("/cloud")

# ------------------------- Config -------------------------
DEFAULT_COMMUNITY = os.getenv("SNMP_COMMUNITY", "public")
FLASK_PORT = int(os.getenv("FLASK_PORT", "5001"))
FLASK_HOST = os.getenv("FLASK_HOST", "0.0.0.0")
TTL_DEFAULT = int(os.getenv("TTL_DEFAULT", "15"))
TIMEOUT_DEFAULT = int(os.getenv("TIMEOUT_DEFAULT", "3"))

# Umbrales XLSX (la UI usa filtro fijo 10% para visual)
XLSX_WARN_PCT = max(1, min(99, int(os.getenv("XLSX_WARN_PCT", "20"))))
XLSX_ALERT_PCT = max(1, min(XLSX_WARN_PCT - 1, int(os.getenv("XLSX_ALERT_PCT", "10"))))
XLSX_WARN_FRAC = XLSX_WARN_PCT / 100.0
XLSX_ALERT_FRAC = XLSX_ALERT_PCT / 100.0

# Auth / sessions
app.config.update(
    SECRET_KEY=os.getenv("SECRET_KEY", "change-me-please"),
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=(os.getenv("SESSION_COOKIE_SECURE", "0") == "1"),
)
AUTH_DB_PATH = os.getenv("AUTH_DB_PATH", "auth.db")

# ------------------------- OIDs ---------------------------
DESC_BASE   = "1.3.6.1.2.1.43.11.1.1.6"   # prtMarkerSuppliesDescription
MAX_BASE    = "1.3.6.1.2.1.43.11.1.1.8"   # prtMarkerSuppliesMaxCapacity
LEVEL_BASE  = "1.3.6.1.2.1.43.11.1.1.9"   # prtMarkerSuppliesLevel
PRT_NAME    = "1.3.6.1.2.1.43.5.1.1.16.1" # prtGeneralPrinterName
SYS_NAME    = "1.3.6.1.2.1.1.5.0"         # sysName
SYS_DESCR   = "1.3.6.1.2.1.1.1.0"         # sysDescr
PRT_LOCATION = "1.3.6.1.2.1.1.6.0"      # sysLocation
PRT_MARKER_LIFECOUNT = "1.3.6.1.2.1.43.10.2.1.4"  # prtMarkerLifeCount
PRT_ALERT_DESC       = "1.3.6.1.2.1.43.18.1.1.8"  # prtAlertDescription

# -------------------- Simple Auth (SQLite) ----------------
def _db():
    conn = sqlite3.connect(AUTH_DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def _init_auth_db():
    with _db() as conn:
        conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );""")

def create_user(username: str, password: str) -> Optional[str]:
    if not username or not password:
        return "Username and password are required."
    if len(password) < 6:
        return "Password must be at least 6 characters."
    try:
        with _db() as conn:
            conn.execute(
                "INSERT INTO users (username, password_hash) VALUES (?, ?)",
                (username.strip(), generate_password_hash(password)),
            )
        return None
    except sqlite3.IntegrityError:
        return "Username already exists."

def verify_user(username: str, password: str) -> bool:
    with _db() as conn:
        cur = conn.execute("SELECT id, password_hash FROM users WHERE username = ?", (username.strip(),))
        row = cur.fetchone()
        if not row:
            return False
        return check_password_hash(row["password_hash"], password)

def find_user_id(username: str) -> Optional[int]:
    with _db() as conn:
        cur = conn.execute("SELECT id FROM users WHERE username = ?", (username.strip(),))
        row = cur.fetchone()
        return int(row["id"]) if row else None

def _update_username(user_id: int, new_username: str) -> Optional[str]:
    try:
        with _db() as conn:
            conn.execute("UPDATE users SET username=? WHERE id=?", (new_username.strip(), user_id))
        return None
    except sqlite3.IntegrityError:
        return "Username already exists."

def _update_password(user_id: int, new_password: str):
    with _db() as conn:
        conn.execute("UPDATE users SET password_hash=? WHERE id=?",
                     (generate_password_hash(new_password), user_id))

def login_required(endpoint_name: str = ""):
    def deco(fn):
        @wraps(fn)
        def _wrap(*args, **kwargs):
            if not session.get("user_id"):
                if request.path.startswith("/api/") or request.headers.get("Accept","").startswith("application/json"):
                    return jsonify({"error": "Unauthorized"}), 401
                nxt = request.full_path if request.query_string else request.path
                return redirect(url_for("login", next=nxt))
            return fn(*args, **kwargs)
        return _wrap
    return deco

def _bootstrap_admin_from_env():
    user = os.getenv("ADMIN_USER")
    pw = os.getenv("ADMIN_PASSWORD")
    if user and pw:
        with _db() as conn:
            row = conn.execute("SELECT 1 FROM users WHERE username=?", (user.strip(),)).fetchone()
            if not row:
                conn.execute("INSERT INTO users (username, password_hash) VALUES (?, ?)",
                             (user.strip(), generate_password_hash(pw)))
                print(f"[bootstrap] Created admin user '{user}' from ENV")

_init_auth_db()
_bootstrap_admin_from_env()

# -------------------- SNMP helper funcs -------------------
def _decode(v: Any) -> Any:
    if isinstance(v, bytes):
        try:
            return v.decode(errors="ignore")
        except Exception:
            return str(v)
    return v

def _snmp_get(host: str, community: str, oid: str, timeout: int) -> str:
    try:
        val = _decode(snmp_get(host, community, oid, timeout=timeout))
        return (str(val) if val is not None else "").strip()
    except Exception:
        return ""

def _snmp_get_name(host: str, community: str, timeout: int) -> str:
    name = _snmp_get(host, community, PRT_NAME, timeout)
    return name or _snmp_get(host, community, SYS_NAME, timeout)

def _guess_hp_model(sys_descr: str, prt_name: str) -> str:
    raw = f"{sys_descr or ''} {prt_name or ''}".strip()
    low = raw.lower()
    if "hp ethernet multi-environment" in low:
        return "HP LaserJet Pro"
    m = re.search(
        r'(?:HP|Hewlett[-\s]?Packard)\s+(?:Color\s+)?'
        r'(?:LaserJet(?:\s+Pro)?|OfficeJet|PageWide|DesignJet|DeskJet)[^\r\n;,]*',
        raw, re.IGNORECASE
    )
    if m:
        model = m.group(0)
        model = model.replace("Hewlett Packard", "HP").replace("Hewlett-Packard", "HP").strip()
        if re.search(r'laserjet\s+pro', model, re.IGNORECASE):
            return "HP LaserJet Pro"
        return model[:80]
    if re.search(r'\bHP\b', raw, re.IGNORECASE):
        return "HP LaserJet Pro"
    return ""

def _snmp_get_model(host: str, community: str, timeout: int) -> str:
    sysd = _snmp_get(host, community, SYS_DESCR, timeout)
    prtn = _snmp_get(host, community, PRT_NAME, timeout)
    cand_lower = f"{(sysd or '').lower()} {(prtn or '').lower()}"
    for token in ("b8155", "b7135", "b415", "c415"):
        if token in cand_lower:
            return token.upper()
    hp_model = _guess_hp_model(sysd, prtn)
    if hp_model:
        return hp_model
    return ""

def _snmp_column_map(host: str, community: str, base_oid: str, timeout: int = TIMEOUT_DEFAULT) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for vb_oid, vb_value in walk(host, community, base_oid, timeout=timeout):
        idx = str(vb_oid).split(".")[-1]
        out[idx] = _decode(vb_value)
    return out

def _safe_pct(level: Any, maxcap: Any) -> float:
    try:
        level = int(level); maxcap = int(maxcap)
    except Exception:
        return -1
    if level < 0 or maxcap <= 0:
        return -1
    return round(100.0 * level / maxcap, 1)

def _get_black_impressions(host: str, community: str, timeout: int) -> int:
    counts: List[int] = []
    try:
        for _, val in walk(host, community, PRT_MARKER_LIFECOUNT, timeout=timeout):
            try: counts.append(int(_decode(val)))
            except Exception: continue
    except Exception:
        return -1
    if not counts: return -1
    pos = [c for c in counts if c >= 0]
    return max(pos) if pos else -1

# --------- Extra para Xerox B7135: leer Fuser/Transfer vÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â­a OID privado ----------
# --------- Extra para Xerox B7135: leer Fuser/Transfer vÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â­a OID privado ----------
def _b7135_r7_r8_status(host: str, community: str, timeout: int) -> Dict[str, str]:
    out = {"fuser": "Unknown", "transfer_roller": "Unknown"}
    try:
        # Recorremos toda la tabla de consumibles Xerox
        base_oid = "1.3.6.1.4.1.253.8.53.13.2.1.6.1.6"
        table = walk(host, community, base_oid, timeout=timeout)

        for oid, val in table:
            oid_str = str(oid)
            val_str = str(val).strip().lower()

            # Fuser R8
            if oid_str.endswith(".8") or "fuser" in oid_str:
                if val_str in ("1", "end", "expired"):
                    out["fuser"] = "Past end of life"
                elif val_str in ("0", "ok", "normal"):
                    out["fuser"] = "OK"

            # Transfer Roller R7
            if oid_str.endswith(".7") or "transfer" in oid_str:
                if val_str in ("1", "end", "expired"):
                    out["transfer_roller"] = "Past end of life"
                elif val_str in ("0", "ok", "normal"):
                    out["transfer_roller"] = "OK"

        # Si no encontramos valores numÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â©ricos, intentar descripciones SNMP
        if out["fuser"] == "Unknown" or out["transfer_roller"] == "Unknown":
            for _, desc in walk(host, community, "1.3.6.1.2.1.43.18.1.1.8", timeout=timeout):
                t = str(desc).lower()
                if "fuser" in t and any(x in t for x in ("end", "replace", "expired", "vida", "fin")):
                    out["fuser"] = "Past end of life"
                if "transfer" in t and any(x in t for x in ("end", "replace", "expired", "vida", "fin")):
                    out["transfer_roller"] = "Past end of life"

        # Defaults
        out["fuser"] = out.get("fuser", "OK")
        out["transfer_roller"] = out.get("transfer_roller", "OK")

    except Exception as e:
        print(f"[DEBUG] Error R7/R8: {e}")
    return out



# -------------------- Categorization ----------------------
def _categorize(desc: str) -> str:
    d = (desc or "").lower()
    if any(k in d for k in ("waste toner","waste container","waste bottle","collection","residuo","residual")):
        return "waste_toner"
    if "r8" in d or "fuser" in d:
        return "fuser"
    if any(k in d for k in ("r7","secondary transfer","second bias","second-bias","sbtr","transfer roll","transfer roller")):
        return "transfer_roller"
    if "clean" in d and "belt" in d:
        return "belt_cleaner"
    if ("belt" in d and "transfer" in d) or "ibt" in d or "r6" in d:
        return "transfer_belt"
    if "imaging" in d or "drum" in d:
        return "drum"
    if ("cartridge" in d or "cartucho" in d) and not any(w in d for w in ("waste","staple","maintenance")):
        return "toner"
    if "toner" in d:
        return "toner"
    return "other"

FRIENDLY_DEFAULT = {
    "toner": "Toner (K)",
    "drum": "Drum / Imaging Unit",
    "fuser": "Fuser R8",
    "transfer_roller": "Transfer Roller R7",
    "waste_toner": "Waste Toner Container",
    "belt_cleaner": "Transfer Belt Cleaner",
    "transfer_belt": "Transfer Belt",
}
FRIENDLY_B8155 = {
    "toner": "Toner",
    "drum": "Drum",
    "waste_toner": "Waste Toner Container",
    "belt_cleaner": "Transfer Belt Cleaner",
    "transfer_roller": "Second Bias Transfer Roll",
}
ALLOWED_B8155 = {"toner", "drum", "waste_toner", "belt_cleaner", "transfer_roller"}

def _friendly_label(category: str, desc: str, model: str) -> str:
    if model == "B8155":
        return FRIENDLY_B8155.get(category, desc or "Other")
    return FRIENDLY_DEFAULT.get(category, desc or "Other")

# -------- B7135: R7/R8 status from alerts --------
ALERT_END_WORDS = [
    "end of life","end-of-life","life end","past of life","past end of life",
    "replace","replacement","change","cambiar","reemplazar",
    "fin de vida","vida util","vida ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Âºtil","near end"
]
# ---------------------- Cache R7/R8 (SNMP alerts) ----------------------
_r7_cache: Dict[str, Tuple[float, Dict[str, str]]] = {}

def _cached_life_status_from_alerts(host: str, community: str, timeout: int) -> Dict[str, str]:
    """
    Igual que _life_status_from_alerts, pero con cachÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â© local de 5 minutos.
    """
    import time
    now = time.time()

    # ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â°ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚ÂÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¹ Si tenemos cache reciente, usarlo
    if host in _r7_cache and (now - _r7_cache[host][0]) < 300:
        print(f"[CACHE] Transfer Roller {host} ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¾ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ {_r7_cache[host][1]}")
        return _r7_cache[host][1]

    # ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â°ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚ÂÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¹ Ejecutar SNMP real
    status = _life_status_from_alerts(host, community, timeout)
    _r7_cache[host] = (time.time(), status)
    print(f"[R7 ALERTS] {host} ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¾ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ {status} (cacheado)")
    return status

def _life_status_from_alerts(host: str, community: str, timeout: int) -> Dict[str, str]:
    status: Dict[str, str] = {}
    try:
        for _, val in walk(host, community, PRT_ALERT_DESC, timeout=timeout):
            text = str(_decode(val)).lower()
            if not text: continue
            if "fuser" in text or "r8" in text:
                status["fuser"] = "END" if any(w in text for w in ALERT_END_WORDS) else "OK"
            if any(k in text for k in ("transfer","roller","r7")):
                status["transfer_roller"] = "END" if any(w in text for w in ALERT_END_WORDS) else "OK"
    except Exception:
        pass
    status.setdefault("fuser","OK"); status.setdefault("transfer_roller","OK")
    return status

# ---- Percent fallback (HP/raw levels) ----
def _fallback_pct_if_raw_percent(category: str, level: Any, maxcap: Any, desc: str = "") -> float:
    p = _safe_pct(level, maxcap)
    if isinstance(p, (int, float)) and p >= 0:
        return p
    if category in ("waste_toner", "toner"):
        try: lv = int(level)
        except Exception: lv = None
        try: mx = int(maxcap)
        except Exception: mx = None
        if lv is not None and (mx is None or mx <= 0) and 0 <= lv <= 100:
            return float(lv)
    return -1

# === C415 helpers ===========================================================
def _color_from_desc(desc: str) -> str:
    d = (desc or "").lower()
    if "cyan" in d or "cian" in d: return "C"
    if "magenta" in d: return "M"
    if "yellow" in d or "amarill" in d: return "Y"
    if "black" in d or "negro" in d or " k" in d or "k toner" in d or " bk " in d or "bk " in d: return "K"
    t0 = (d.split() or [""])[0]
    if t0 in ("c","cyan","cian"): return "C"
    if t0 in ("m","magenta"): return "M"
    if t0 in ("y","yellow","amarillo","amarilla"): return "Y"
    if t0 in ("k","black","bk","negro"): return "K"
    return ""

def _extract_c415(items: List[Dict[str, Any]]) -> Dict[str, Any]:
    out: Dict[str, Any] = {
        "toner": {}, "drum": {},
        "waste_toner": None, "belt_cleaner": None,
        "transfer_belt": None, "transfer_roller": None
    }
    def best(existing: Any, newp: Any) -> Any:
        en = existing if isinstance(existing, (int, float)) else -1
        np = newp if isinstance(newp, (int, float)) else -1
        return np if np > en else existing

    for it in items:
        cat = it.get("category")
        desc = it.get("description") or ""
        p = it.get("percent") if isinstance(it.get("percent"), (int, float)) else -1
        if cat == "toner":
            col = _color_from_desc(desc)
            if col: out["toner"][col] = best(out["toner"].get(col, -1), p)
        elif cat == "drum":
            col = _color_from_desc(desc)
            if col: out["drum"][col] = best(out["drum"].get(col, -1), p)
        elif cat in ("waste_toner","belt_cleaner","transfer_belt","transfer_roller"):
            if (not isinstance(p,(int,float))) or p < 0:
                p = _fallback_pct_if_raw_percent(cat, it.get("level"), it.get("max_capacity"), desc)
            out[cat] = best(out[cat], p)
    return out
# --------- B7135: Lectura Fuser y captura visual del panel ---------
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time, re, os 
from datetime import datetime

# ---------------------- Fuser R8 (B7135) con cachÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â© ----------------------
_fuser_cache: Dict[str, Tuple[float, str]] = {}

def _b7135_fuser_status(host: str):
    """
    Usa Selenium headless para obtener el estado del Fuser R8 en la Xerox B7135.
    Devuelve 'OK', 'Past end of life' o 'Unknown'.
    Con cachÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â© de 1 hora para evitar ejecuciones repetitivas lentas.
    """
    import time
    now = time.time()

    # ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â°ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚ÂÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¹ Si ya hay un valor en cachÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â© reciente (menos de 1h), usarlo
    if host in _fuser_cache and (now - _fuser_cache[host][0]) < 3600:
        cached_status = _fuser_cache[host][1]
        print(f"[CACHE] Fuser {host} ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¾ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ {cached_status}")
        return cached_status

    status = "Unknown"
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.chrome.service import Service
        from webdriver_manager.chrome import ChromeDriverManager

        url = f"http://{host}/"
        chrome_options = Options()
        chrome_options.add_argument("--headless=new")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--log-level=3")
        chrome_options.add_argument("--window-size=1920,1080")

        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        driver.set_page_load_timeout(20)
        driver.get(url)
        time.sleep(6)  # esperar carga de pÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¡gina
        page_source = driver.page_source.lower()
        driver.quit()

        if "fuser r8" in page_source or "fuser" in page_source:
            if any(word in page_source for word in ("end", "replace", "expired", "vida", "fin")):
                status = "Past end of life"
            else:
                status = "OK"

    except Exception as e:
        print(f"[Fuser Selenium Error] {host}: {e}")

    # ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â°ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚ÂÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¹ Guardar en cachÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â© el resultado con timestamp
    _fuser_cache[host] = (time.time(), status)
    print(f"[B7135 Fuser] {host} ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¾ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ {status} (cacheado)")

    return status


# --------- B7135: Lectura Fuser/Transfer R7 vÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â­a interfaz web (con cachÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â©) ---------
_fuser_tr_cache: Dict[str, Tuple[float, Dict[str, str]]] = {}

def _b7135_web_status(host: str) -> Dict[str, str]:
    """
    Usa Selenium headless para obtener el estado del Fuser R8 y Transfer Roller R7
    desde la interfaz web del Xerox B7135.
    Devuelve {'fuser': 'OK'/'Past end of life', 'transfer_roller': 'OK'/'Past end of life'}.
    Mantiene un cachÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â© de 1 hora para evitar abrir el navegador en cada consulta.
    """
    import time
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager

    now = time.time()
    if host in _fuser_tr_cache and (now - _fuser_tr_cache[host][0]) < 3600:
        return _fuser_tr_cache[host][1]

    result = {"fuser": "Unknown", "transfer_roller": "Unknown"}
    try:
        url = f"http://{host}/"
        opts = Options()
        opts.add_argument("--headless=new")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-gpu")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--log-level=3")
        opts.add_argument("--window-size=1920,1080")

        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)
        driver.set_page_load_timeout(25)
        driver.get(url)

        # ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â°ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚ÂÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¹ Espera dinÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¡mica: hasta que aparezca texto "fuser" o "transfer roller"
        import time
        start = time.time()
        page = ""
        while time.time() - start < 12:  # mÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¡ximo 12 s
            page = driver.page_source.lower()
            if "fuser" in page or "transfer roller" in page or "r7" in page:
                break
            time.sleep(1.2)  # revisar cada segundo
        driver.quit()

        if "fuser r8" in page or "fuser" in page:
            if any(word in page for word in ("past end of life", "end of life", "replace", "expired", "fin de vida")):
                result["fuser"] = "Past end of life"
            else:
                result["fuser"] = "OK"

        if "transfer roller r7" in page or "transfer roller" in page or "r7" in page:
            if any(word in page for word in ("past end of life", "end of life", "replace", "expired", "fin de vida")):
                result["transfer_roller"] = "Past end of life"
            else:
                result["transfer_roller"] = "OK"

    except Exception as e:
        print(f"[B7135 Selenium Error] {host}: {e}")


    _fuser_tr_cache[host] = (time.time(), result)
    return result


def _b7135_take_snapshot(host: str):
    """
    Toma una captura visual del panel del B7135 y la guarda en /snapshots/.
    """
    try:
        chrome_options = Options()
        chrome_options.add_argument("--headless=new")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--window-size=1920,1080")
        chrome_options.add_argument("--log-level=3")

        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        driver.set_page_load_timeout(25)
        driver.get(f"http://{host}/")
        time.sleep(8)

        out_dir = os.path.join(os.getcwd(), "snapshots")
        os.makedirs(out_dir, exist_ok=True)

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"B7135_{host.replace('.', '-')}_{timestamp}.png"
        out_path = os.path.join(out_dir, filename)
        driver.save_screenshot(out_path)
        driver.quit()

        print(f"[Snapshot B7135] Guardada: {out_path}")
        return filename

    except Exception as e:
        print(f"[Snapshot Error] {e}")
        return None

# -------------------- Core SNMP fetch ---------------------
def fetch_supplies_generic(ip: str, community: str = DEFAULT_COMMUNITY, timeout: int = TIMEOUT_DEFAULT) -> Dict[str, Any]:
    name = _snmp_get_name(ip, community, timeout)
    model = _snmp_get_model(ip, community, timeout)
    location = _snmp_get(ip, community, PRT_LOCATION, timeout)  # Address / Business in sysLocation

    descs  = _snmp_column_map(ip, community, DESC_BASE,  timeout=timeout)
    maxs   = _snmp_column_map(ip, community, MAX_BASE,   timeout=timeout)
    levels = _snmp_column_map(ip, community, LEVEL_BASE, timeout=timeout)
    black_impr = _get_black_impressions(ip, community, timeout)

    # --- Ajuste especial B7135: Fuser R8 + Transfer Roller R7 ---
    life_status = {}
    if model == "B7135":
        try:
            # Prioridad: leer desde el panel web (cachÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â© 1h) para Fuser R8 y Transfer Roller R7
            web_status = _b7135_web_status(ip)
            fuser_status = web_status.get("fuser", "Unknown")
            tr_status = web_status.get("transfer_roller", "Unknown")
        except Exception as e:
            print(f"[B7135] Error leyendo panel web: {e}")
            fuser_status = "Unknown"
            tr_status = "Unknown"

        # Si alguno sigue Unknown, intentar hÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â­brido por SNMP (alertas + OID privado)
        if fuser_status == "Unknown" or tr_status == "Unknown":
            try:
                hybrid = _b7135_r7_r8_status(ip, community, timeout)
            except Exception as e:
                print(f"[B7135] Error hÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â­brido R7/R8: {e}")
                hybrid = {}
            if fuser_status == "Unknown":
                fuser_status = hybrid.get("fuser", "Unknown")
            if tr_status == "Unknown":
                tr_status = hybrid.get("transfer_roller", "Unknown")

        life_status = {
            "fuser": "END" if "end" in str(fuser_status).lower() else "OK",
            "transfer_roller": "END" if "end" in str(tr_status).lower() else "OK"
        }

        print(f"[B7135 FINAL] {ip} -> Fuser={fuser_status}, Transfer Roller={tr_status}")

    # --- Generar lista de items base SNMP ---
    items: List[Dict[str, Any]] = []
    for idx in sorted(set(descs) | set(maxs) | set(levels), key=lambda x: int(x)):
        desc = descs.get(idx)
        cat  = _categorize(desc or "")
        if model == "B8155" and cat not in ALLOWED_B8155:
            continue
        maxc = maxs.get(idx)
        lev = levels.get(idx)
        pct = _fallback_pct_if_raw_percent(cat, lev, maxc, desc or "")

        entry = {
            "index": int(idx),
            "description": desc,
            "max_capacity": maxc,
            "level": lev,
            "percent": pct,
            "category": cat,
            "label": _friendly_label(cat, desc or "Other", model),
        }

        # Aplicar estados de vida ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Âºtil al B7135
        if model == "B7135" and cat in ("fuser", "transfer_roller"):
            st = life_status.get(cat, "OK")
            entry["status"] = "Past end of life" if st.lower() in ("end", "past end of life") else "OK"
            entry["percent"] = -1

        items.append(entry)

    # --- Modelo HP (rescate) ---
    if (model or "").upper().startswith("HP"):
        toners = [it for it in items if it.get("category") == "toner"]
        has_pct = any(isinstance(it.get("percent"), (int, float)) and it["percent"] >= 0 for it in toners)
        if not toners or not has_pct:
            cand = None
            kws = ("cartridge", "cartucho", "black", "negro", "toner")
            for it in items:
                d = (it.get("description") or "").lower()
                if any(k in d for k in kws):
                    cand = it
                    break
            if cand is None and len(items) == 1:
                cand = items[0]
            if cand is not None:
                lv = cand.get("level")
                mx = cand.get("max_capacity")
                p = _fallback_pct_if_raw_percent("toner", lv, mx, cand.get("description") or "")
                if (not isinstance(p, (int, float))) or p < 0:
                    try:
                        ilv = int(lv)
                        if 0 <= ilv <= 100:
                            p = float(ilv)
                    except Exception:
                        pass
                cand["category"] = "toner"
                cand["label"] = FRIENDLY_DEFAULT["toner"]
                cand["percent"] = p if isinstance(p, (int, float)) else -1

    # --- B8155 split & pick ---
    if model == "B8155":
        toners = [it for it in items if it.get("category") == "toner"]
        if len(toners) >= 2:
            candidate = None
            for it in toners:
                d = (it.get("description") or "").lower()
                if any(k in d for k in ("waste", "container", "bottle", "collection")):
                    candidate = it
                    break
            if candidate is None:
                candidate = max(
                    toners,
                    key=lambda it: it.get("percent")
                    if isinstance(it.get("percent"), (int, float))
                    else -1,
                )
            candidate["category"] = "waste_toner"
            candidate["label"] = FRIENDLY_B8155["waste_toner"]
        toners_left = [it for it in items if it.get("category") == "toner"]
        if toners_left:
            chosen = None
            for it in toners_left:
                dd = (it.get("description") or "").lower()
                if any(k in dd for k in ("toner cartridge", "black toner", "k toner", "bk toner")):
                    chosen = it
                    break
            if chosen is None:
                chosen = max(
                    toners_left,
                    key=lambda it: it.get("percent")
                    if isinstance(it.get("percent"), (int, float))
                    else -1,
                )
            for it in toners_left:
                if it is not chosen:
                    it["category"] = "other"

    # --- Ordenar y extraer C415 ---
    items.sort(
        key=lambda it: (
            it["category"] == "other",
            -(it["percent"] if isinstance(it["percent"], (int, float)) else -1),
        )
    )

    c415_map: Dict[str, Any] = {}
    if model == "C415":
        c415_map = _extract_c415(items)

    return {
        "ip": ip,
        "printer_name": name,
        "model": model,
        "location": location,
        "black_impressions": black_impr,
        "items": items,
        "c415": c415_map if model == "C415" else {},
    }


# -------------------------- Cache -------------------------
CACHE: Dict[Tuple[str, str, int], Tuple[float, Dict[str, Any]]] = {}
def get_cached(ip: str, community: str, timeout: int, ttl: int) -> Dict[str, Any]:
    key = (ip, community, timeout)
    now = time.time()
    hit = CACHE.get(key)
    if hit and hit[0] > now:
        return hit[1]
    data = fetch_supplies_generic(ip, community, timeout)
    CACHE[key] = (now + max(1, ttl), data)
    return data

# ----------------------- AUTH PAGES -----------------------
def _lang() -> str:
    lang = session.get("lang", "en")
    return lang if lang in {"en", "es"} else "en"


def _txt(en: str, es: str) -> str:
    return es if _lang() == "es" else en


def _login_form(next_url: str = "/cloud") -> str:
    username = _txt("Username", "Usuario")
    password = _txt("Password", "Contrasena")
    username_ph = _txt("Enter your username", "Ingresa tu usuario")
    password_ph = _txt("Enter your password", "Ingresa tu contrasena")
    show_pw = _txt("Show", "Ver")
    show_label = _txt("Show password", "Mostrar contrasena")
    sign_in = _txt("Sign in", "Iniciar sesion")
    return f"""
      <form class="login-form" method="post" action="/login?next={htmlmod.escape(next_url)}" novalidate>
        <div class="field">
          <label for="username">{htmlmod.escape(username)}</label>
          <input id="username" name="username" autocomplete="username" placeholder="{htmlmod.escape(username_ph)}" autofocus required>
        </div>
        <div class="field">
          <label for="password">{htmlmod.escape(password)}</label>
          <div class="password-field">
            <input id="password" name="password" type="password" autocomplete="current-password" placeholder="{htmlmod.escape(password_ph)}" required>
            <button id="pwToggle" class="toggle-pass" type="button" aria-label="{htmlmod.escape(show_label)}" data-show="{htmlmod.escape(show_pw)}" data-hide="{htmlmod.escape(_txt('Hide', 'Ocultar'))}">{htmlmod.escape(show_pw)}</button>
          </div>
        </div>
        <button class="login-btn" type="submit">{htmlmod.escape(sign_in)}</button>
      </form>
    """


def _auth_base_html(body: str, title: str = "Printers Supplies", error: str = "") -> str:
    lang = _lang()
    err = f'<div class="login-error" role="alert">{htmlmod.escape(error)}</div>' if error else ""
    overview = _txt("System overview", "Resumen del sistema")
    eyebrow = _txt("Xerox Supplies Monitor", "Monitor de suministros Xerox")
    hero_title = _txt("Fast visibility into your printer fleet.", "Visibilidad rapida de tu flota de impresoras.")
    hero_copy = _txt("Track supplies, models, and alerts from one operational support dashboard.", "Consulta suministros, modelos y alertas desde un panel operativo unico.")
    ready = _txt("Ready", "Listo")
    active = _txt("Active", "Activo")
    kicker = _txt("Supplies monitor", "Monitor de suministros")
    welcome = _txt("Welcome back", "Bienvenido de nuevo")
    intro = _txt("Enter your credentials to access the monitoring dashboard.", "Ingresa tus credenciales para acceder al panel de monitoreo.")
    language = _txt("Language", "Idioma")
    return f"""<!doctype html>
<html lang="{lang}">
<head>
  <meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <title>{htmlmod.escape(title)} - Supplies</title>
  <style>
    *{{box-sizing:border-box}}
    body{{margin:0;min-height:100vh;color:#0f172a;font-family:ui-sans-serif,system-ui,-apple-system,Segoe UI,Roboto,Arial;background:linear-gradient(90deg,rgba(15,23,42,.94),rgba(15,23,42,.70)),radial-gradient(circle at 18% 20%,rgba(14,165,233,.34),transparent 22rem),linear-gradient(135deg,#0f172a 0%,#1e3a5f 54%,#eef6ff 54%,#f8fafc 100%);}}
    .login-shell{{min-height:100vh;display:grid;grid-template-columns:minmax(0,1fr) minmax(360px,440px);align-items:center;gap:56px;max-width:1120px;margin:0 auto;padding:40px;}}
    .login-hero{{color:#fff}} .login-hero-card{{max-width:560px}} .login-eyebrow{{margin:0 0 16px;color:#93c5fd;font-size:13px;font-weight:800;letter-spacing:.08em;text-transform:uppercase}}
    .login-hero h2{{margin:0;font-size:48px;line-height:1.05}} .login-hero p{{margin:18px 0 0;max-width:520px;color:#dbeafe;font-size:17px;line-height:1.6}}
    .login-stats{{display:flex;flex-wrap:wrap;gap:10px;margin-top:32px}} .login-stats span{{display:inline-flex;gap:7px;align-items:center;padding:9px 12px;color:#e2e8f0;background:rgba(255,255,255,.08);border:1px solid rgba(226,232,240,.16);border-radius:8px;font-size:13px}} .login-stats strong{{color:#fff}}
    .login-panel{{width:100%;padding:34px;background:#fff;border:1px solid rgba(148,163,184,.26);border-radius:8px;box-shadow:0 28px 80px rgba(15,23,42,.26);}}
    .login-brand{{display:flex;align-items:center;gap:14px;margin-bottom:28px}} .login-mark{{width:48px;height:48px;display:grid;place-items:center;color:#2563eb;background:#e0f2fe;border:1px solid #bae6fd;border-radius:8px;flex:0 0 auto}}
    .login-kicker{{margin:0 0 3px;color:#64748b;font-size:12px;font-weight:700;text-transform:uppercase}} .login-brand h1{{margin:0;font-size:24px;line-height:1.15}}
    .login-copy{{margin-bottom:22px}} .login-copy h2{{margin:0;font-size:23px;line-height:1.2}} .login-copy p{{margin:8px 0 0;color:#64748b;font-size:14px;line-height:1.55}}
    .login-error{{margin-bottom:18px;padding:11px 12px;color:#991b1b;background:#fef2f2;border:1px solid #fecaca;border-radius:8px;font-size:14px;font-weight:600}}
    .login-form{{display:grid;gap:16px}} .field label{{display:block;margin-bottom:7px;color:#334155;font-size:13px;font-weight:700}}
    .field input{{width:100%;min-height:48px;padding:11px 13px;color:#0f172a;background:#fff;border:1px solid #cbd5e1;border-radius:8px;font:inherit;font-size:14px;outline:none}} .field input:focus{{border-color:#2563eb;box-shadow:0 0 0 4px rgba(37,99,235,.14)}}
    .password-field{{position:relative}} .password-field input{{padding-right:84px}} .toggle-pass{{position:absolute;right:8px;top:50%;min-width:62px;min-height:32px;padding:0 10px;color:#2563eb;background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;cursor:pointer;font:inherit;font-size:13px;font-weight:700;transform:translateY(-50%)}}
    .login-btn{{width:100%;min-height:48px;margin-top:4px;color:#fff;background:#2563eb;border:0;border-radius:999px;cursor:pointer;font:inherit;font-size:15px;font-weight:800;box-shadow:0 12px 24px rgba(37,99,235,.24)}} .login-btn:hover{{background:#1d4ed8}}
    .language-switcher{{display:flex;align-items:center;justify-content:center;gap:8px;margin-top:20px;color:#64748b;font-size:13px}} .language-switcher a{{padding:5px 9px;color:#2563eb;border:1px solid #bfdbfe;border-radius:8px;text-decoration:none;font-weight:800}} .language-switcher a.active{{color:#fff;background:#2563eb;border-color:#2563eb}}
    @media(max-width:860px){{body{{background:radial-gradient(circle at top left,rgba(14,165,233,.18),transparent 30rem),linear-gradient(135deg,#f8fafc 0%,#eef6ff 100%)}}.login-shell{{grid-template-columns:1fr;max-width:480px;gap:24px;padding:24px}}.login-hero{{color:#0f172a}}.login-hero h2{{font-size:30px}}.login-hero p{{color:#475569;font-size:15px}}.login-stats{{display:none}}}}
    @media(max-width:480px){{.login-shell{{padding:16px}}.login-hero{{display:none}}.login-panel{{padding:24px}}.login-brand h1{{font-size:21px}}}}
  </style>
</head>
<body>
  <main class="login-shell" aria-labelledby="login-title">
    <section class="login-hero" aria-label="{htmlmod.escape(overview)}">
      <div class="login-hero-card">
        <p class="login-eyebrow">{htmlmod.escape(eyebrow)}</p>
        <h2>{htmlmod.escape(hero_title)}</h2>
        <p>{htmlmod.escape(hero_copy)}</p>
        <div class="login-stats" aria-hidden="true"><span><strong>B7135</strong> {htmlmod.escape(ready)}</span><span><strong>B415</strong> Online</span><span><strong>SNMP</strong> {htmlmod.escape(active)}</span></div>
      </div>
    </section>
    <section class="login-panel">
      <div class="login-brand"><div class="login-mark" aria-hidden="true"><svg viewBox="0 0 24 24" width="28" height="28" fill="none"><path d="M7 8V4h10v4" stroke="currentColor" stroke-width="1.9" stroke-linecap="round" stroke-linejoin="round"/><rect x="4" y="8" width="16" height="8" rx="2" stroke="currentColor" stroke-width="1.9"/><path d="M7 16v4h10v-4" stroke="currentColor" stroke-width="1.9" stroke-linecap="round" stroke-linejoin="round"/><circle cx="17" cy="12" r="1" fill="#10b981"/></svg></div><div><p class="login-kicker">{htmlmod.escape(kicker)}</p><h1 id="login-title">{htmlmod.escape(title)}</h1></div></div>
      <div class="login-copy"><h2>{htmlmod.escape(welcome)}</h2><p>{htmlmod.escape(intro)}</p></div>
      {err}
      {body}
      <nav class="language-switcher" aria-label="{htmlmod.escape(language)}"><span>{htmlmod.escape(language)}</span><a class="{'active' if lang == 'en' else ''}" href="/language/en">EN</a><a class="{'active' if lang == 'es' else ''}" href="/language/es">ES</a></nav>
    </section>
  </main>
  <script>
    (function(){{const input=document.getElementById('password');const btn=document.getElementById('pwToggle');if(!input||!btn)return;btn.addEventListener('click',function(){{const show=input.type==='password';input.type=show?'text':'password';btn.textContent=show?btn.dataset.hide:btn.dataset.show;}});}})();
  </script>
</body>
</html>"""


@app.get("/login")
def login():
    if session.get("user_id"):
        return redirect("/cloud")
    return _auth_base_html(_login_form(request.args.get("next", "/cloud")), "Printers Supplies")


@app.post("/login")
def login_post():
    username = (request.form.get("username") or "").strip()
    password = request.form.get("password") or ""
    next_url = request.args.get("next", "/cloud")
    if not verify_user(username, password):
        return _auth_base_html(_login_form(next_url), "Printers Supplies", _txt("Invalid username or password.", "Usuario o contrasena incorrectos."))
    uid = find_user_id(username)
    session["user_id"] = uid
    session["username"] = username
    return redirect(next_url or "/cloud")


@app.get("/logout")
def logout():
    lang = session.get("lang")
    session.clear()
    if lang:
        session["lang"] = lang
    return redirect(url_for("login"))


@app.get("/language/<lang>")
def set_language(lang):
    if lang in {"en", "es"}:
        session["lang"] = lang
    next_url = request.args.get("next", "")
    if next_url.startswith("/") and not next_url.startswith("//"):
        return redirect(next_url)
    return redirect(request.referrer or url_for("login"))
# ----------------------- ACCOUNT (Edit user) -----------------------
def _account_html(form_html: str, message: str = "", is_error: bool = False) -> str:
    note = ""
    if message:
        note_class = "alert error" if is_error else "alert success"
        note = f'<div class="{note_class}">{htmlmod.escape(message)}</div>'
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Account - Printers Supplies</title>
  <style>
    :root{{
      --page:#eef5ff; --panel:#ffffff; --ink:#061936; --muted:#52627a;
      --line:#cfddf0; --soft:#f7fbff; --blue:#2f66e8; --blue-2:#86b1ff;
      --good:#047857; --good-bg:#ecfdf5; --bad:#b42318; --bad-bg:#fff2f1;
    }}
    *{{box-sizing:border-box}}
    body{{
      margin:0; min-height:100vh; color:var(--ink);
      font-family:ui-sans-serif,system-ui,-apple-system,Segoe UI,Roboto,Arial,sans-serif;
      background:
        radial-gradient(circle at top left, rgba(91,141,239,.18), transparent 34rem),
        linear-gradient(180deg,#f8fbff 0%,var(--page) 100%);
    }}
    .shell{{max-width:980px;margin:0 auto;padding:28px 22px 56px}}
    .topbar{{
      display:flex;align-items:center;justify-content:space-between;gap:18px;
      background:rgba(255,255,255,.86);border:1px solid var(--line);border-radius:22px;
      padding:18px 20px;box-shadow:0 18px 42px rgba(31,52,92,.08);
    }}
    .brand{{display:flex;align-items:center;gap:14px}}
    .brand-mark{{
      width:46px;height:46px;border-radius:14px;border:1px solid #b9d5ff;
      background:linear-gradient(180deg,#f4f9ff,#e9f2ff);display:grid;place-items:center;
      color:var(--blue);font-weight:900;font-size:17px;
    }}
    .brand-title{{font-size:24px;line-height:1;font-weight:900;letter-spacing:0;color:#72a0f7}}
    .brand-sub{{margin-top:6px;color:#3d5068;font-size:13px;font-weight:700}}
    .nav{{display:flex;align-items:center;gap:14px;flex-wrap:wrap;color:#34445c;font-size:14px}}
    .nav a{{color:#124bd3;text-decoration:none;font-weight:800}}
    .nav .sep{{color:#a4b1c4;font-weight:800}}
    .card{{
      margin-top:18px;background:rgba(255,255,255,.94);border:1px solid var(--line);
      border-radius:22px;box-shadow:0 24px 56px rgba(31,52,92,.10);overflow:hidden;
    }}
    .card-head{{display:flex;align-items:center;gap:14px;padding:24px 24px 18px;border-bottom:1px solid #dbe6f5}}
    .section-icon{{
      width:42px;height:42px;border-radius:12px;border:1px solid #b8d4ff;background:#eff6ff;
      color:var(--blue);display:grid;place-items:center;font-weight:900;
    }}
    h1{{margin:0;font-size:22px;line-height:1.15}}
    p.muted{{color:var(--muted);margin:6px 0 0;font-size:14px}}
    form{{padding:22px 24px 24px}}
    label{{display:block;font-size:12px;color:#52627a;margin:0 0 8px;font-weight:900;text-transform:uppercase;letter-spacing:.06em}}
    input{{
      width:100%;height:44px;border:1px solid #d4dfef;border-radius:12px;
      padding:10px 12px;font-size:14px;background:#fff;color:var(--ink);outline:none;
      transition:border-color .15s ease, box-shadow .15s ease;
    }}
    input:focus{{border-color:#86b1ff;box-shadow:0 0 0 4px rgba(47,102,232,.12)}}
    .btn{{
      margin-top:18px;background:var(--blue);color:#fff;border:0;border-radius:12px;
      padding:12px 18px;font-weight:900;cursor:pointer;box-shadow:0 12px 28px rgba(47,102,232,.22);
    }}
    .btn:hover{{filter:brightness(.96)}}
    .row{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-top:14px}}
    .row.single{{grid-template-columns:1fr}}
    .grow{{min-width:0}}
    a{{color:var(--blue);text-decoration:none}}
    .pw{{position:relative}}
    .pw input{{padding-right:76px}}
    .pwbtn{{
      position:absolute;right:7px;top:50%;transform:translateY(-50%);
      height:32px;min-width:58px;border:1px solid #bed5ff;border-radius:9px;background:#f2f7ff;
      cursor:pointer;color:#124bd3;font-size:12px;font-weight:900;
    }}
    .pwbtn:focus{{outline:3px solid rgba(47,102,232,.18)}}
    .alert{{margin:0 24px 24px;padding:12px 14px;border-radius:12px;font-size:14px;font-weight:700}}
    .alert.error{{color:var(--bad);background:var(--bad-bg);border:1px solid #ffb4ad}}
    .alert.success{{color:var(--good);background:var(--good-bg);border:1px solid #99f6c5}}
    @media (max-width:720px){{
      .shell{{padding:16px 12px 36px}}
      .topbar,.card-head{{align-items:flex-start;flex-direction:column}}
      .row{{grid-template-columns:1fr}}
      .brand-title{{font-size:22px}}
    }}
  </style>
</head>
<body>
  <div class="shell">
    <div class="topbar">
      <div class="brand">
        <div class="brand-mark">PS</div>
        <div>
          <div class="brand-title">Printers Supplies</div>
          <div class="brand-sub">Account settings and access control</div>
        </div>
      </div>
      <div class="nav">
        <a href="/cloud">Back to dashboard</a>
        <span class="sep">|</span>
        <a href="/logout">Logout</a>
      </div>
    </div>
    <div class="card">
      <div class="card-head">
        <div class="section-icon">ID</div>
        <div>
          <h1>Account</h1>
          <p class="muted">Update your username or password for this dashboard.</p>
        </div>
      </div>
      {form_html}
      {note}
    </div>
  </div>
  <script>
  const toggles = [
    ['current_password','pwToggleCur'],
    ['new_password','pwToggleNew'],
    ['confirm_password','pwToggleCnf']
  ];
  for (const [inputId, btnId] of toggles) {{
    const input = document.getElementById(inputId);
    const btn = document.getElementById(btnId);
    if (input && btn) {{
      btn.addEventListener('click', ()=>{{ 
        const show = input.type === 'password';
        input.type = show ? 'text' : 'password';
        btn.textContent = show ? 'Hide' : 'Show';
        btn.title = show ? 'Hide' : 'Show';
        btn.setAttribute('aria-label', show ? 'Hide password' : 'Show password');
      }});
    }}
  }}
</script>

</body>
</html>"""

@app.get("/account")
@login_required("account_get")
def account_get():
    username = session.get("username","")
    return _account_html(_account_form_prefill(username))

@app.post("/account")
@login_required("account_post")
def account_post():
    current_user = session.get("username","")
    uid = session.get("user_id")
    if not uid:
        return redirect(url_for("login"))

    username_new = (request.form.get("username") or "").strip()
    current_pw   = request.form.get("current_password") or ""
    new_pw       = request.form.get("new_password") or ""
    confirm_pw   = request.form.get("confirm_password") or ""

    if not verify_user(current_user, current_pw):
        return _account_html(_account_form_prefill(username_new), "Current password is incorrect.", True)

    if username_new and username_new != current_user:
        err = _update_username(uid, username_new)
        if err:
            return _account_html(_account_form_prefill(username_new), err, True)
        session["username"] = username_new

    if new_pw or confirm_pw:
        if len(new_pw) < 6:
            return _account_html(_account_form_prefill(username_new), "New password must be at least 6 characters.", True)
        if new_pw != confirm_pw:
            return _account_html(_account_form_prefill(username_new), "Passwords do not match.", True)
        _update_password(uid, new_pw)

    return _account_html(_account_form_prefill(username_new or current_user), "Changes saved successfully.", False)

def _account_form_prefill(username_value: str) -> str:
    return f"""
      <form method="post" action="/account">
        <div class="row single">
          <div class="grow">
            <label>Username</label>
            <input name="username" value="{htmlmod.escape(username_value)}" placeholder="your.name">
          </div>
        </div>
        <div class="row">
          <div class="grow">
            <label>Current password</label>
            <div class="pw">
              <input id="current_password" name="current_password" type="password" placeholder="Current password" required>
              <button type="button" id="pwToggleCur" class="pwbtn" title="Show" aria-label="Show password">Show</button>
            </div>
          </div>
        </div>
        <div class="row">
          <div class="grow">
            <label>New password (optional)</label>
            <div class="pw">
              <input id="new_password" name="new_password" type="password" placeholder="min. 6 characters">
              <button type="button" id="pwToggleNew" class="pwbtn" title="Show" aria-label="Show password">Show</button>
            </div>
          </div>
          <div class="grow">
            <label>Confirm new password</label>
            <div class="pw">
              <input id="confirm_password" name="confirm_password" type="password" placeholder="repeat new password">
              <button type="button" id="pwToggleCnf" class="pwbtn" title="Show" aria-label="Show password">Show</button>
            </div>
          </div>
        </div>
        <button class="btn" type="submit">Save changes</button>
      </form>
    """

# ----------------------- Web (UI) -------------------------
def _parse_int(value: Optional[str], default: int) -> int:
    try:
        v = int(str(value))
        if v <= 0: return default
        return v
    except Exception:
        return default

@app.get("/")
@login_required("home")
def home():
    return redirect("/cloud")
    username = session.get("username") or ""
    signed_in_as = _txt("Signed in as", "Sesion iniciada como")
    edit_account = _txt("Edit Account", "Editar cuenta")
    logout_text = _txt("Logout", "Cerrar sesion")
    user_menu = f'''
      <nav class="user-menu">
        <span class="um-muted">{htmlmod.escape(signed_in_as)}</span> <strong>{htmlmod.escape(username)}</strong>
        <span class="um-sep">|</span><a href="/language/en">EN</a>
        <span class="um-sep">|</span><a href="/language/es">ES</a>
        <span class="um-sep">|</span><a href="/account">{htmlmod.escape(edit_account)}</a>
        <span class="um-sep">|</span><a href="/logout">{htmlmod.escape(logout_text)}</a>
      </nav>'''

    html = r'''
<!doctype html>
<html lang="__LANG__">
<head>
  <meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Printers Supplies</title>
  <style>
    :root{
      --bg:#f6f7fb; --card:#ffffff; --text:#0f172a; --muted:#64748b;
      --border:#e5e7eb; --accent:#2563eb; --accent-2:#10b981;
      --okbg:#e7f7ee; --ok:#2e7d32; --badbg:#fdecea; --bad:#b3261e;
      --zebra:#fbfcff; --hover:#eef6ff; --bar:#eef2f7;
      --toast-bg:#0f172a; --toast-fg:#ffffff;
    }
    *{box-sizing:border-box}
    body{
      margin:0; background:var(--bg); color:var(--text);
      font-family: ui-sans-serif, system-ui, -apple-system, "Segoe UI", Roboto, Arial, "Noto Sans", "Helvetica Neue", sans-serif;
    }
    .container{max-width:min(95vw,1600px); margin:28px auto; padding:0 16px;}

    /* Top bar */
    .topbar{display:flex; align-items:center; justify-content:space-between; margin-bottom:18px}
    .title{display:flex; align-items:center; gap:12px}
    .title-text{
      font-size:28px; font-weight:900; letter-spacing:.3px; margin:0;
      background:linear-gradient(90deg,#60a5fa,#2563eb); -webkit-background-clip:text; background-clip:text; color:transparent;
      text-shadow:0 1px 0 rgba(255,255,255,.35);
    }
    .printer-ico{ width:32px; height:32px; }

    /* User menu ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã¢â‚¬Â¦Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“ bonito */
    .user-menu{
      display:flex; align-items:center; gap:10px;
      font-size:14.5px; color:#0f172a; letter-spacing:.2px; font-weight:700;
    }
    .user-menu .um-muted{ color:#6b7280; font-weight:600; letter-spacing:.3px; }
    .user-menu strong{
      font-weight:900;
      background:linear-gradient(90deg,#0ea5e9,#2563eb);
      -webkit-background-clip:text; background-clip:text; color:transparent;
    }
    .user-menu .um-sep{ color:#c7d2fe; font-weight:900; margin:0 2px; }
    .user-menu a{
      color:#1d4ed8; text-decoration:none; font-weight:800;
      padding:2px 6px; border-radius:10px; border-bottom:2px solid transparent;
      transition: all .18s ease;
    }
    .user-menu a:hover{ background:#eef2ff; border-bottom-color:#93c5fd; transform: translateY(-.5px); }

    /* Printer link (no purple) */
    a.printer-link{ color:#1d4ed8; font-weight:800; text-decoration:none; border-bottom:1px dashed rgba(29,78,216,.35); }
    a.printer-link:hover{ background:#eef6ff; border-radius:6px; padding:0 2px; }


    /* Cards / inputs */
    .card{background:var(--card); border:1px solid var(--border); border-radius:16px; padding:16px; box-shadow:0 6px 22px rgba(2,6,23,.06)}
    .sticky-card{ position: sticky; top: 12px; z-index: 30; }
    .toolbar{display:flex; gap:12px; align-items:end; flex-wrap:wrap}
    label{font-size:12px; color:var(--muted); letter-spacing:.6px; text-transform:uppercase; font-weight:800}
    input, textarea{
      width:100%; border:1px solid var(--border); background:#fff; color:#0f172a;
      border-radius:14px; padding:10px 12px; font-size:14px; outline:0; transition:all .15s ease;
      box-shadow: 0 1px 0 rgba(255,255,255,.7) inset;
    }
    input:focus, textarea:focus{border-color:var(--accent); box-shadow:0 0 0 3px rgba(37,99,235,.15)}
    textarea{height:110px; resize:vertical}
    /* Ajuste especÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â­fico del textarea de IPs */
    textarea#ips{height:110px;}

    /* ===== Mobile fix ===== */
    @media (max-width: 720px){
      .container{padding:0 12px;}
      .w-200, .w-280, .w-420{width:100% !important;}
      .row{gap:12px;}
      .toolbar{align-items:stretch;}
      textarea#ips{
        height:72px !important;
        min-height:72px !important;
        max-height:140px;
      }
    }

    .stack{display:flex; flex-direction:column; gap:6px}
    .row{display:flex; gap:16px; flex-wrap:wrap}
    .w-200{width:200px} .w-280{width:280px} .w-420{width:420px}

    .btn{
      appearance:none; border:0; border-radius:12px; padding:10px 14px; font-weight:800; letter-spacing:.2px;
      cursor:pointer; transition:transform .06s ease, box-shadow .2s ease, background .2s ease, color .2s ease;
      box-shadow: 0 8px 20px rgba(37,99,235,.12);
    }
    .btn:active{transform:translateY(1px)}
    .btn-primary{background:linear-gradient(135deg, #2563eb, #0ea5e9); color:#fff; box-shadow:0 10px 24px rgba(37,99,235,.25)}
    .btn-secondary{background:transparent; color:var(--accent); border:1px solid var(--accent)}
    #spinner{display:none; width:18px; height:18px; border:3px solid #cbd5e1; border-top-color:var(--accent); border-radius:50%;
             animation:spin .8s linear infinite; margin-left:8px}
    @keyframes spin{to{transform:rotate(360deg)}}

    /* Results bar (tÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â­tulo izq, filtros der) */
    .resultsbar{ display:flex; align-items:center; justify-content:space-between; gap:12px; margin:20px 0 8px; }
    .section-title{ font-size:20px; font-weight:900; letter-spacing:.3px; color:#0b1220; margin:0; }
    .right-tools{ display:flex; align-items:center; gap:10px; }

    /* Checkbox + select de umbral */
    .check-inline{
      display:flex; align-items:center; gap:10px; user-select:none;
      background:#ffffff; border:1px solid #dbe6ff; padding:6px 10px; border-radius:12px;
      box-shadow:0 4px 12px rgba(37,99,235,.08);
    }
    .check-inline span{ font-size:13.5px; font-weight:800; color:#0f172a; letter-spacing:.3px; }
    .check-inline input{ width:16px; height:16px; }
    .threshold{
      height:34px; border:1px solid #dbe6ff; border-radius:10px; padding:0 10px; font-weight:800; color:#0f172a;
      background:#fff; outline:0;
    }

    /* Tabla */
    .tableWrap{margin-top:16px; background:var(--card); border:1px solid var(--border);
      border-radius:16px; overflow:auto; box-shadow:0 6px 22px rgba(2,6,23,.06)}
    table{width:100%; border-collapse:separate; border-spacing:0}
    thead th{
      position:sticky; top:0; z-index:1; font-weight:700; font-size:13px;
      background:linear-gradient(#f8fafc,#eef2f7); color:#0f172a; text-align:left;
      padding:12px 10px; border-bottom:1px solid var(--border); cursor:pointer; user-select:none
    }
    thead th .sort{font-weight:400; opacity:.6; margin-left:6px}
    tbody td{padding:12px 10px; border-bottom:1px solid var(--border); font-size:14.25px; vertical-align:middle}
    tbody tr:nth-child(odd) td{background:var(--zebra)}
    tbody tr:hover td{background:var(--hover)}
    .nowrap{white-space:nowrap}
    .pill{display:inline-block;padding:3px 10px;border-radius:999px;border:1px solid #99c;background:#eef;font-weight:600;font-size:12px}
    .pill.ok{background:var(--okbg);border-color:#7cc79f;color:#2e7d32}
    .pill.bad{background:var(--badbg);border-color:#f5b5b0;color:#b3261e}
    .bar{height:10px;background:var(--bar);border-radius:999px;overflow:hidden}
    .bar>div{height:100%;border-radius:inherit;transition:width .35s ease}
    .muted{color:var(--muted)}
    .c415-card{margin-top:12px; background:var(--card); border:1px solid var(--border);
      border-radius:16px; padding:16px; box-shadow:0 6px 22px rgba(2,6,23,.06)}
    .kv{display:grid; grid-template-columns:repeat(auto-fit,minmax(220px,1fr)); gap:12px; margin-top:6px}
    .metric{background:#f9fafb; border:1px solid #eef2ff; border-radius:12px; padding:10px 12px;}
    .metric .label{font-size:12px; color:var(--muted); margin-bottom:4px}
    .err{color:#b00; font-weight:600}
    .empty{border:1px dashed var(--border); border-radius:16px; padding:24px; color:var(--muted); text-align:center}
    .badge{display:inline-block;padding:2px 8px;border-radius:999px;font-weight:800;font-size:11px}
    .badge.alert{background:#fef2f2;border:1px solid #fecaca;color:#b91c1c}

    /* Toast con icono y cerrar */
    .toast-container{ position:fixed; right:16px; bottom:16px; display:flex; flex-direction:column; gap:10px; z-index:9999; }
    .toast{
      background:var(--toast-bg); color:var(--toast-fg); padding:10px 12px; border-radius:12px;
      box-shadow:0 10px 30px rgba(0,0,0,.25); min-width:260px; opacity:0; transform:translateY(8px);
      transition:opacity .2s ease, transform .2s ease; font-weight:700; letter-spacing:.2px;
    }
    .toast.show{ opacity:1; transform:translateY(0); }
    .toast.success{ background:#065f46; }
    .toast.warn{ background:#92400e; }
    .toast.error{ background:#7f1d1d; }
    .toast .row{ display:flex; align-items:center; gap:10px; }
    .toast .icon{ width:18px; height:18px; flex:0 0 18px; }
    .toast .msg{ flex:1; }
    .toast .close{
      background:transparent; color:#fff; border:0; font-size:16px; line-height:1; cursor:pointer;
      padding:2px 6px; border-radius:8px;
    }
    .toast .close:hover{ background:rgba(255,255,255,.12); }

    /* Modern dashboard refresh */
    body{
      background:
        radial-gradient(circle at top left, rgba(14,165,233,.14), transparent 32rem),
        linear-gradient(180deg,#f7faff 0%,#f3f6fb 44%,#eef2f8 100%);
    }
    .container{max-width:min(92vw,1500px); margin:30px auto 56px; padding:0}
    .topbar{
      position:relative; align-items:flex-start; gap:18px; margin-bottom:18px; padding:0 2px 14px;
      border-bottom:1px solid rgba(148,163,184,.18);
    }
    .title-text{font-size:30px; line-height:1; letter-spacing:0}
    .printer-ico{width:34px;height:34px}
    .user-menu{
      flex-wrap:wrap; justify-content:flex-end; gap:8px; padding-top:4px;
      font-size:14px; color:#475569;
    }
    .user-menu a{
      padding:7px 10px; border:1px solid transparent; border-radius:10px; background:transparent;
      color:#1d4ed8; border-bottom:0;
    }
    .user-menu a:hover{background:#eff6ff; border-color:#bfdbfe; transform:none}
    .user-menu .um-sep{color:#cbd5e1; margin:0}

    .card.sticky-card{
      position:relative; top:auto; overflow:hidden; padding:0; border:1px solid rgba(148,163,184,.28);
      border-radius:22px; background:rgba(255,255,255,.92); box-shadow:0 20px 60px rgba(15,23,42,.08);
    }
    .card.sticky-card::before{
      content:""; display:block; height:4px;
      background:linear-gradient(90deg,#2563eb,#0ea5e9,#10b981);
    }
    .toolbar{
      display:grid; grid-template-columns:minmax(126px,160px) minmax(360px,1fr) auto auto;
      gap:16px; align-items:end; padding:18px;
    }
    .switch{
      min-height:110px; display:flex; flex-direction:column; justify-content:center; align-items:flex-start;
      gap:10px; padding:14px; border:1px solid #dbeafe; border-radius:16px; background:#f8fbff;
      color:#0f172a; font-weight:900;
    }
    .switch input{width:18px;height:18px; accent-color:#2563eb}
    .switch span{font-size:12px; letter-spacing:.08em; text-transform:uppercase; color:#475569}
    .stack label{font-size:11px; letter-spacing:.1em; color:#475569}
    textarea#ips{height:110px; min-height:110px; border-radius:16px; line-height:1.5}
    input, textarea, .threshold{
      border-color:#dbe3ef; box-shadow:0 1px 2px rgba(15,23,42,.04);
    }
    .btn{height:42px; border-radius:12px; white-space:nowrap}
    .btn-primary{
      background:linear-gradient(135deg,#2563eb,#0284c7);
      box-shadow:0 12px 26px rgba(37,99,235,.24);
    }
    .btn-secondary{
      background:#ffffff; border:1px solid #bfdbfe; color:#1d4ed8;
      box-shadow:0 8px 20px rgba(37,99,235,.08);
    }
    .btn-secondary:hover{background:#eff6ff}

    .dashboard-metrics{
      display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:12px; margin:16px 0 20px;
    }
    .metric-card{
      background:#fff; border:1px solid rgba(148,163,184,.28); border-radius:18px; padding:14px 16px;
      box-shadow:0 10px 30px rgba(15,23,42,.05);
    }
    .metric-card span{
      display:block; margin-bottom:7px; color:#64748b; font-size:11px; font-weight:900;
      letter-spacing:.09em; text-transform:uppercase;
    }
    .metric-card strong{display:block; color:#0f172a; font-size:26px; line-height:1; letter-spacing:0}
    .metric-card small{display:block; margin-top:7px; color:#94a3b8; font-weight:700}

    .resultsbar{margin:22px 0 10px}
    .section-title{font-size:22px; letter-spacing:0}
    .right-tools{gap:10px}
    .check-inline,.threshold{
      height:38px; border-radius:12px; border-color:#dbeafe; box-shadow:0 8px 24px rgba(37,99,235,.07);
    }
    .check-inline input{accent-color:#2563eb}
    .check-inline span{font-size:12px; letter-spacing:.04em; text-transform:uppercase}
    #filter{height:42px; border-radius:14px}
    .empty{
      padding:34px; background:rgba(255,255,255,.55); border-color:#d9e2ef; color:#64748b;
    }
    .tableWrap{
      border-radius:18px; box-shadow:0 18px 44px rgba(15,23,42,.08); border-color:rgba(148,163,184,.30);
    }
    thead th{
      padding:14px 12px; background:#f8fafc; color:#334155; font-size:12px; letter-spacing:.04em;
      text-transform:uppercase;
    }
    tbody td{padding:14px 12px}
    tbody tr:hover td{background:#f0f7ff}
    .bar{height:9px; background:#e2e8f0}
    .pill,.badge{letter-spacing:.03em}

    @media (max-width: 980px){
      .toolbar{grid-template-columns:1fr; align-items:stretch}
      .switch{min-height:auto; flex-direction:row; align-items:center}
      .dashboard-metrics{grid-template-columns:repeat(2,minmax(0,1fr))}
      .topbar{flex-direction:column}
      .user-menu{justify-content:flex-start}
    }
    @media (max-width: 560px){
      .container{max-width:none; margin:18px 12px 40px}
      .dashboard-metrics{grid-template-columns:1fr}
      .resultsbar{align-items:flex-start; flex-direction:column}
      .right-tools{width:100%; justify-content:space-between}
      .title-text{font-size:24px}
    }

    /* Dashboard v2 polish */
    .topbar{
      padding:20px 22px; margin-bottom:16px; border:1px solid rgba(148,163,184,.22);
      border-radius:24px; background:linear-gradient(135deg,rgba(255,255,255,.96),rgba(239,246,255,.82));
      box-shadow:0 18px 50px rgba(15,23,42,.07);
    }
    .brand-copy{display:flex; flex-direction:column; gap:4px}
    .title-subtitle{margin:0; color:#64748b; font-size:13px; font-weight:700}
    .title{
      min-width:260px;
    }
    .user-menu{
      padding:7px; border:1px solid rgba(191,219,254,.7); border-radius:16px;
      background:rgba(255,255,255,.76); box-shadow:0 8px 24px rgba(37,99,235,.06);
    }
    .user-menu .um-muted{font-size:13px}
    .user-menu a{border-radius:11px}

    .dashboard-metrics{margin:0 0 14px}
    .metric-card{
      position:relative; overflow:hidden; min-height:102px;
      background:linear-gradient(180deg,#fff,#fbfdff);
    }
    .metric-card::after{
      content:""; position:absolute; right:-28px; top:-28px; width:86px; height:86px;
      border-radius:999px; background:rgba(37,99,235,.08);
    }
    .metric-card:nth-child(2)::after{background:rgba(239,68,68,.08)}
    .metric-card:nth-child(3)::after{background:rgba(16,185,129,.09)}
    .metric-card:nth-child(4)::after{background:rgba(14,165,233,.10)}

    .card.sticky-card{margin-bottom:20px}
    .query-card-head{
      display:flex; align-items:center; justify-content:space-between; gap:14px;
      padding:18px 18px 10px;
      border-bottom:1px solid rgba(226,232,240,.72);
    }
    .query-card-title{display:flex; align-items:center; gap:12px}
    .query-icon{
      width:42px; height:42px; display:grid; place-items:center; color:#2563eb;
      background:#eff6ff; border:1px solid #bfdbfe; border-radius:14px;
    }
    .query-card-title h2{margin:0; font-size:18px; line-height:1.2}
    .query-card-title p{display:none}
    .query-status-pill{
      display:inline-flex; align-items:center; gap:8px; padding:8px 11px; border-radius:999px;
      color:#0369a1; background:#e0f2fe; border:1px solid #bae6fd; font-size:12px; font-weight:900;
      text-transform:uppercase; letter-spacing:.04em;
    }
    .query-status-pill::before{
      content:""; width:8px; height:8px; border-radius:999px; background:#10b981;
      box-shadow:0 0 0 4px rgba(16,185,129,.14);
    }
    .toolbar{
      grid-template-columns:minmax(128px,168px) minmax(390px,1fr) auto auto;
      padding:16px 18px 18px;
    }
    .switch{
      background:linear-gradient(180deg,#f8fbff,#eef6ff); border-color:#bfdbfe;
    }
    .switch span{color:#1e3a8a}
    .btn-primary,.btn-secondary{min-width:128px}
    .btn-primary:hover{transform:translateY(-1px); box-shadow:0 16px 34px rgba(37,99,235,.28)}

    .resultsbar{
      padding:0 2px;
    }
    .section-heading{display:flex; flex-direction:column; gap:0}
    .section-heading p{display:none}
    .search-card{
      margin-bottom:12px; padding:12px; border:1px solid rgba(148,163,184,.22);
      border-radius:18px; background:rgba(255,255,255,.88); box-shadow:0 10px 30px rgba(15,23,42,.04);
    }
    .empty{
      min-height:112px; display:grid; place-items:center; background:
        linear-gradient(180deg,rgba(255,255,255,.76),rgba(248,250,252,.76));
      font-size:14px;
      font-weight:650;
    }
    .empty::before{
      content:""; width:38px; height:38px; margin-bottom:8px; display:block;
      border-radius:14px; background:linear-gradient(135deg,#dbeafe,#e0f2fe);
      box-shadow:inset 0 0 0 1px #bfdbfe;
    }
    @media (max-width: 980px){
      .query-card-head{align-items:flex-start; flex-direction:column}
      .query-status-pill{align-self:flex-start}
      .toolbar{grid-template-columns:1fr}
    }

    /* Visual icon pass */
    .toolbar{
      grid-template-columns:minmax(390px,1fr) auto auto;
    }
    .switch{
      display:none;
    }
    #singleInputs{
      display:none !important;
    }
    #listInputs{
      display:flex !important;
    }
    .metric-card{
      display:grid;
      grid-template-columns:46px minmax(0,1fr);
      column-gap:13px;
      align-items:center;
    }
    .metric-icon{
      width:46px;
      height:46px;
      display:grid;
      place-items:center;
      color:#2563eb;
      background:#eff6ff;
      border:1px solid #bfdbfe;
      border-radius:14px;
      grid-row:1 / span 3;
      position:relative;
      z-index:1;
    }
    .metric-card:nth-child(2) .metric-icon{
      color:#dc2626;
      background:#fef2f2;
      border-color:#fecaca;
    }
    .metric-card:nth-child(3) .metric-icon{
      color:#059669;
      background:#ecfdf5;
      border-color:#bbf7d0;
    }
    .metric-card:nth-child(4) .metric-icon{
      color:#0284c7;
      background:#f0f9ff;
      border-color:#bae6fd;
    }
    .btn{
      display:inline-flex;
      align-items:center;
      justify-content:center;
      gap:8px;
    }
    .btn svg,
    .query-icon svg,
    .metric-icon svg{
      flex:0 0 auto;
    }
    .empty{
      gap:8px;
    }
    .empty::before{
      content:"";
      width:46px;
      height:46px;
      margin-bottom:0;
      background:#eff6ff;
      border:1px solid #bfdbfe;
      border-radius:14px;
      box-shadow:none;
    }
    @media (max-width: 980px){
      .toolbar{grid-template-columns:1fr}
      .btn{width:100%}
    }

    /* Inventory table typography */
    .tableWrap{
      background:#fff;
      border-radius:18px;
      overflow:auto;
    }
    table{
      min-width:980px;
    }
    thead th{
      padding:11px 14px;
      color:#5f6678;
      background:#f8fafc;
      font-size:11px;
      font-weight:900;
      letter-spacing:.07em;
    }
    tbody td{
      padding:14px 14px;
      font-size:13px;
      border-bottom:1px solid #edf1f7;
    }
    .printer-cell{
      min-width:250px;
    }
    .printer-name{
      display:inline-block;
      color:#030712;
      font-size:16px;
      font-weight:900;
      line-height:1.15;
      text-decoration:none;
      border-bottom:0;
    }
    .printer-name:hover{
      color:#1d4ed8;
      background:transparent;
      padding:0;
    }
    .printer-meta{
      margin-top:5px;
      color:#64748b;
      font-family:ui-monospace, SFMono-Regular, Menlo, Consolas, monospace;
      font-size:12px;
      line-height:1.35;
      letter-spacing:.01em;
    }
    .model-chip{
      display:inline-flex;
      align-items:center;
      min-height:22px;
      padding:3px 9px;
      color:#334155;
      background:#f1f5f9;
      border:1px solid #e2e8f0;
      border-radius:999px;
      font-size:11px;
      font-weight:900;
    }
    .supply-meter{
      display:flex;
      flex-direction:column;
      align-items:stretch;
      gap:5px;
      width:150px;
      min-width:128px;
      max-width:150px;
    }
    .supply-meter::before,
    .supply-meter::after{
      content:none !important;
      display:none !important;
    }
    .supply-meter .bar{
      width:100%;
      height:5px;
      background:#eef2f7;
      border-radius:999px;
    }
    .supply-meter .bar > div{
      border-radius:999px;
    }
    .supply-percent{
      color:#16a34a;
      font-size:11px;
      line-height:1;
      font-weight:900;
      text-align:left;
    }
    .supply-percent.warn{
      color:#f59e0b;
    }
    .supply-percent.low{
      color:#dc2626;
    }
    .pill{
      padding:4px 10px;
      font-size:11px;
      font-weight:900;
      text-transform:uppercase;
    }
    .pill.ok{
      background:#dcfce7;
      border-color:#bbf7d0;
      color:#15803d;
    }
    .pill.bad{
      background:#fee2e2;
      border-color:#fecaca;
      color:#dc2626;
    }
    .badge.alert{
      display:none;
    }

    /* Orientation helpers */
    .section-nav{
      position:sticky;
      top:10px;
      z-index:35;
      display:flex;
      align-items:center;
      justify-content:space-between;
      gap:12px;
      margin:0 0 14px;
      padding:9px;
      background:rgba(255,255,255,.86);
      border:1px solid rgba(191,219,254,.72);
      border-radius:16px;
      box-shadow:0 14px 36px rgba(15,23,42,.07);
      backdrop-filter:blur(12px);
    }
    .section-nav-links{
      display:flex;
      flex-wrap:wrap;
      gap:8px;
    }
    .section-nav a,
    .section-nav button{
      min-height:34px;
      display:inline-flex;
      align-items:center;
      gap:7px;
      padding:0 12px;
      border-radius:11px;
      border:1px solid transparent;
      background:transparent;
      color:#475569;
      text-decoration:none;
      font:inherit;
      font-size:13px;
      font-weight:850;
      cursor:pointer;
    }
    .section-nav a:hover,
    .section-nav button:hover{
      background:#eff6ff;
      color:#1d4ed8;
      border-color:#bfdbfe;
    }
    .section-nav a.active{
      background:#2563eb;
      color:#fff;
      box-shadow:0 8px 20px rgba(37,99,235,.22);
    }
    .section-nav svg{
      flex:0 0 auto;
    }
    .back-top{
      position:fixed;
      right:22px;
      bottom:22px;
      z-index:80;
      width:44px;
      height:44px;
      display:grid;
      place-items:center;
      color:#fff;
      background:#2563eb;
      border:0;
      border-radius:999px;
      box-shadow:0 18px 38px rgba(37,99,235,.34);
      cursor:pointer;
      opacity:0;
      pointer-events:none;
      transform:translateY(8px);
      transition:opacity .18s ease, transform .18s ease;
    }
    .back-top.show{
      opacity:1;
      pointer-events:auto;
      transform:translateY(0);
    }
    html{
      scroll-behavior:smooth;
    }
    #lookup,
    #summary,
    #results{
      scroll-margin-top:86px;
    }
    @media (max-width: 720px){
      .section-nav{
        position:relative;
        top:auto;
        align-items:flex-start;
        flex-direction:column;
      }
      .section-nav-actions{
        width:100%;
        display:flex;
      }
      .section-nav button{
        flex:1;
        justify-content:center;
      }
    }

  </style>
</head>
<body>
  <div class="container">

    <div class="topbar">
      <div class="title">
        <svg class="printer-ico" viewBox="0 0 24 24" fill="none">
          <path d="M7 8V4h10v4" stroke="#2563eb" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
          <rect x="3" y="8" width="18" height="8" rx="2" stroke="#0ea5e9" stroke-width="2"/>
          <path d="M7 16v4h10v-4" stroke="#2563eb" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
          <circle cx="17.5" cy="12" r="1" fill="#10b981"/>
        </svg>
        <div class="brand-copy">
          <h1 class="title-text">Printers Supplies</h1>
          <p class="title-subtitle">Supply monitoring and fleet health overview</p>
        </div>
      </div>
      __USER_MENU__
    </div>

    <nav class="section-nav" aria-label="Dashboard sections">
      <div class="section-nav-links">
        <a class="active" href="#lookup" data-section-link="lookup">
          <svg viewBox="0 0 24 24" width="16" height="16" fill="none" aria-hidden="true"><path d="M10.5 18a7.5 7.5 0 1 1 5.3-2.2L21 21" stroke="currentColor" stroke-width="2" stroke-linecap="round"/></svg>
          Lookup
        </a>
        <a href="#summary" data-section-link="summary">
          <svg viewBox="0 0 24 24" width="16" height="16" fill="none" aria-hidden="true"><path d="M4 19V5m4 12v2m4-9v9m4-6v6m4-13v13" stroke="currentColor" stroke-width="2" stroke-linecap="round"/></svg>
          Summary
        </a>
        <a href="#results" data-section-link="results">
          <svg viewBox="0 0 24 24" width="16" height="16" fill="none" aria-hidden="true"><path d="M4 6h16M4 12h16M4 18h10" stroke="currentColor" stroke-width="2" stroke-linecap="round"/></svg>
          Results
        </a>
      </div>
      <div class="section-nav-actions">
        <button id="btnClear" type="button">
          <svg viewBox="0 0 24 24" width="16" height="16" fill="none" aria-hidden="true"><path d="M4 7h16M10 11v6m4-6v6M6 7l1 14h10l1-14M9 7V4h6v3" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/></svg>
          Clear
        </button>
      </div>
    </nav>

    <!-- QUERY CARD -->
    <div id="lookup" class="card sticky-card">
      <div class="query-card-head">
        <div class="query-card-title">
          <div class="query-icon" aria-hidden="true">
            <svg viewBox="0 0 24 24" width="22" height="22" fill="none">
              <path d="M10.5 18a7.5 7.5 0 1 1 5.3-2.2L21 21" stroke="currentColor" stroke-width="2" stroke-linecap="round"/>
              <path d="M7.5 10.5h6M7.5 13h4" stroke="currentColor" stroke-width="2" stroke-linecap="round"/>
            </svg>
          </div>
          <div>
            <h2>Supply Lookup</h2>
          </div>
        </div>
        <div class="query-status-pill">Ready</div>
      </div>
      <form id="f" class="toolbar" autocomplete="off">
        <input type="checkbox" id="listmode" checked hidden aria-hidden="true">

        <div id="singleInputs" class="row" style="display:none">
          <div class="stack w-280">
            <label>IP</label>
            <input id="ip" placeholder="192.168.1.100">
          </div>
        </div>

        <div id="listInputs" class="row">
          <div class="stack w-420">
            <label>IPs (one per line or comma-separated)</label>
            <textarea id="ips" placeholder="192.168.1.100
192.168.1.101"></textarea>
          </div>
        </div>

        <button id="btnQ" class="btn btn-primary">
          <svg viewBox="0 0 24 24" width="18" height="18" fill="none" aria-hidden="true">
            <path d="M10.5 18a7.5 7.5 0 1 1 5.3-2.2L21 21" stroke="currentColor" stroke-width="2" stroke-linecap="round"/>
            <path d="M7 10h7M7 13h4" stroke="currentColor" stroke-width="2" stroke-linecap="round"/>
          </svg>
          Query
        </button>
        <button id="btnXLSX" type="button" class="btn btn-secondary">
          <svg viewBox="0 0 24 24" width="18" height="18" fill="none" aria-hidden="true">
            <path d="M12 3v12m0 0 4-4m-4 4-4-4M5 21h14" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
          </svg>
          Export to Excel
        </button>
      </form>
    </div>

    <section id="summary" class="dashboard-metrics" aria-label="Dashboard summary">
      <div class="metric-card">
        <div class="metric-icon" aria-hidden="true">
          <svg viewBox="0 0 24 24" width="22" height="22" fill="none">
            <path d="M7 8V4h10v4M7 17v3h10v-3M5 17h14a2 2 0 0 0 2-2v-5a2 2 0 0 0-2-2H5a2 2 0 0 0-2 2v5a2 2 0 0 0 2 2Z" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
          </svg>
        </div>
        <span>Total printers</span>
        <strong id="metricTotal">0</strong>
        <small>Total in current query</small>
      </div>
      <div class="metric-card">
        <div class="metric-icon" aria-hidden="true">
          <svg viewBox="0 0 24 24" width="22" height="22" fill="none">
            <path d="M12 9v4m0 4h.01M10.3 4.3 2.8 17.2A2 2 0 0 0 4.5 20h15a2 2 0 0 0 1.7-2.8L13.7 4.3a2 2 0 0 0-3.4 0Z" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
          </svg>
        </div>
        <span>Low supplies</span>
        <strong id="metricLow">0</strong>
        <small>Supplies requiring attention</small>
      </div>
      <div class="metric-card">
        <div class="metric-icon" aria-hidden="true">
          <svg viewBox="0 0 24 24" width="22" height="22" fill="none">
            <path d="M4 7h16M7 7v10a2 2 0 0 0 2 2h6a2 2 0 0 0 2-2V7M9 7V5a2 2 0 0 1 2-2h2a2 2 0 0 1 2 2v2" stroke="currentColor" stroke-width="2" stroke-linecap="round"/>
          </svg>
        </div>
        <span>Model types</span>
        <strong id="metricModels">0</strong>
        <small>Unique printer models detected</small>
      </div>
      <div class="metric-card">
        <div class="metric-icon" aria-hidden="true">
          <svg viewBox="0 0 24 24" width="22" height="22" fill="none">
            <path d="M12 8v5l3 2M21 12a9 9 0 1 1-3-6.7" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
            <path d="M21 4v5h-5" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
          </svg>
        </div>
        <span>Last update</span>
        <strong id="metricUpdated">--</strong>
        <small>Local browser time</small>
      </div>
    </section>

    <!-- Results bar (con umbral seleccionable) -->
    <div id="results" class="resultsbar">
      <div class="section-heading">
        <h2 class="section-title">Results</h2>
      </div>
      <div class="right-tools">
        <label class="check-inline">
          <input id="onlyLow" type="checkbox">
          <span>Show only below</span>
        </label>
        <select id="threshold" class="threshold">
          <option value="10" selected>10%</option>
          <option value="15">15%</option>
          <option value="20">20%</option>
        </select>
      </div>
    </div>

    <div class="row search-card">
      <div class="stack" style="flex:1">
        <label>Search</label>
        <input id="filter" placeholder="Filter by IP, Printer or Model">
      </div>
    </div>

    <div id="out" class="empty">No data yet. Run a query to see results.</div>
    <div id="extras"></div>
  </div>

  <button id="backTop" class="back-top" type="button" aria-label="Back to top">
    <svg viewBox="0 0 24 24" width="20" height="20" fill="none" aria-hidden="true"><path d="M12 19V5m0 0-6 6m6-6 6 6" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/></svg>
  </button>

<!-- Printer Info Modal -->
<div id="printerModal" role="dialog" aria-modal="true" aria-labelledby="pmTitle" style="display:none;position:fixed;inset:0;z-index:9998;">
  <div id="pmBackdrop" style="position:absolute;inset:0;background:rgba(2,6,23,.55);backdrop-filter: blur(2px);"></div>
  <div style="position:relative;z-index:9999;display:grid;place-items:center;min-height:100%;">
    <div style="width:min(92vw,560px);background:#fff;border:1px solid #e5e7eb;border-radius:18px;box-shadow:0 25px 80px rgba(2,6,23,.35);padding:18px 18px 14px;">
      <div style="display:flex;align-items:center;justify-content:space-between;gap:10px;">
        <div>
          <div id="pmTitle" style="font-size:20px;font-weight:900;margin:0;color:#0f172a;">Printer Info</div>
          <div id="pmSubtitle" style="margin-top:2px;color:#64748b;font-weight:700;font-size:13px;"></div>
        </div>
        <button id="pmCloseX" type="button" aria-label="Close" style="border:0;background:#f1f5f9;border-radius:12px;width:36px;height:36px;cursor:pointer;font-size:18px;font-weight:900;color:#0f172a;">x</button>
      </div>

      <div style="margin-top:12px;display:grid;gap:12px;">
        <div>
          <div style="font-size:11px;letter-spacing:.14em;text-transform:uppercase;color:#94a3b8;font-weight:900;">Address</div>
          <div id="pmAddress" style="margin-top:4px;font-size:14.5px;font-weight:800;color:#0f172a;word-break:break-word;"></div>
        </div>
        <div>
          <div style="font-size:11px;letter-spacing:.14em;text-transform:uppercase;color:#94a3b8;font-weight:900;">Business</div>
          <div id="pmBusiness" style="margin-top:4px;font-size:14.5px;font-weight:800;color:#0f172a;word-break:break-word;"></div>
        </div>
      </div>

      <button id="pmCloseBtn" type="button" class="btn btn-primary" style="margin-top:14px;width:100%;">Close</button>
    </div>
  </div>
</div>

  <!-- Toast container -->
  <div class="toast-container" id="toasts"></div>

  <script>
    // ====== Utils ======
    function pctColor(p){ if(p<0) return '#cbd5e1'; if(p>=50) return '#16a34a'; if(p>=20) return '#f59e0b'; return '#ef4444'; }

    // ====== Toast con icono + cerrar ======
    function showToast(msg, type='warn', ms=2600){
      const icons = {
        success: '<svg class="icon" viewBox="0 0 24 24" fill="none"><path d="M20 7L9 18l-5-5" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/></svg>',
        warn:    '<svg class="icon" viewBox="0 0 24 24" fill="none"><path d="M12 9v4m0 4h.01M10.29 3.86l-8.49 14.7A2 2 0 003.53 21h16.94a2 2 0 001.73-3l-8.49-14.7a2 2 0 00-3.42 0z" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/></svg>',
        error:   '<svg class="icon" viewBox="0 0 24 24" fill="none"><path d="M15 9l-6 6m0-6l6 6" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/></svg>'
      };
      const box = document.getElementById('toasts');
      const el = document.createElement('div');
      el.className = 'toast ' + (type||'');
      el.innerHTML = `<div class="row">
          ${icons[type] || icons.warn}
          <div class="msg">${msg}</div>
          <button class="close" aria-label="Close">x</button>
        </div>`;
      box.appendChild(el);
      requestAnimationFrame(()=> el.classList.add('show'));
      const kill = ()=>{ el.classList.remove('show'); setTimeout(()=> el.remove(), 180); };
      el.querySelector('.close').addEventListener('click', kill);
      setTimeout(kill, ms);
    }

    // ====== Estado / Controles ======
    const state = { rows:[], sortKey:'ip', sortDir:'asc', filter:'', onlyLow:false, threshold:10 };

    const listmode = document.getElementById('listmode');
    const singleInputs = document.getElementById('singleInputs');
    const listInputs = document.getElementById('listInputs');
    if (listmode) {
      listmode.checked = true;
      singleInputs.style.display = 'none';
      listInputs.style.display = 'flex';
      listmode.addEventListener('change', ()=>{
        const on = listmode.checked;
        singleInputs.style.display = on ? 'none' : 'flex';
        listInputs.style.display   = on ? 'flex' : 'none';
      });
    }

    const cbLow = document.getElementById('onlyLow');
    const selTh = document.getElementById('threshold');
    cbLow.addEventListener('change', ()=>{ state.onlyLow = !!cbLow.checked; renderTable(); });
    selTh.addEventListener('change', ()=>{
      const v = parseInt(selTh.value, 10);
      state.threshold = isNaN(v) ? 10 : v;
      renderTable();
    });

    function setMetric(id, value){
      const el = document.getElementById(id);
      if(el) el.textContent = value;
    }

    function updateMetrics(allRows){
      const rows = allRows || (state.rows || []).map(r => r.raw);
      const models = new Set(rows.map(d => String(d.model || '').trim()).filter(Boolean));
      const low = rows.filter(rowIsBelowThreshold).length;
      setMetric('metricTotal', rows.length);
      setMetric('metricLow', low);
      setMetric('metricModels', models.size);
      setMetric('metricUpdated', rows.length ? new Date().toLocaleTimeString([], {hour:'2-digit', minute:'2-digit'}) : '--');
    }

    // ====== Helpers de render ======
    function valueHTML(item){
      if(!item) return '';
      if(item.status){
        const good = String(item.status).toLowerCase()==='ok';
        return `<span class="pill ${good?'ok':'bad'}">${item.status}</span>`;
      }
      const p = (typeof item.percent==='number' && item.percent>=0) ? item.percent : -1;
      if(p<0) return '';
      const w = Math.max(0, Math.min(100, p));
      const statusClass = p < 20 ? ' low' : (p < 50 ? ' warn' : '');
      return `<div class="supply-meter">
                <div class="bar"><div style="width:${w}%;background:${pctColor(p)}"></div></div>
                <span class="supply-percent${statusClass}">${p.toFixed(0)}%</span>
              </div>`;
    }
    function pickBest(items, category){
      const arr = items.filter(it => it.category === category);
      if(!arr.length) return null;
      return arr.slice().sort((a,b)=>{
        const pa = (typeof a.percent==='number')?a.percent:-1;
        const pb = (typeof b.percent==='number')?b.percent:-1;
        return pb-pa;
      })[0];
    }

    const COLS = [
      {key:'printer', label:'Printer Information'},
      {key:'model', label:'Status'},
      {key:'black', label:'Black Impressions'},
      {key:'toner', label:'Toner'},
      {key:'drum', label:'Drum'},
      {key:'tr_r7', label:'Transfer R7'},
      {key:'fuser', label:'Fuser R8'},
      {key:'waste', label:'Waste Toner'},
      {key:'beltcl', label:'Belt Cleaner'},
      {key:'sbtr', label:'Bias Transfer'}
    ];

    function headerHTML(sortKey, sortDir){
      return `<thead><tr>${
        COLS.map(c=>{
          const arrow = (sortKey===c.key) ? (sortDir==='asc'?'^':'v') : '';
          return `<th data-key="${c.key}">${c.label}<span class="sort">${arrow}</span></th>`;
        }).join('')
      }</tr></thead>`;
    }

    function pivotRowHTML(d){
      const items = d.items || [];
      const model = (d.model||'').toUpperCase();
      if(model==='C415') return '';

      const toner   = pickBest(items,'toner');
      const drum    = pickBest(items,'drum');
      const tr      = pickBest(items,'transfer_roller');
      const fuser   = pickBest(items,'fuser');
      const waste   = pickBest(items,'waste_toner');
      const beltcl  = pickBest(items,'belt_cleaner');

      let trR7 = null, sbtr = null;
      if(tr){ if(model==='B8155') sbtr = tr; else trR7 = tr; }

      let bi = 'N/A';
      if(typeof d.black_impressions === 'number' && d.black_impressions >= 0){
        try{ bi = Number(d.black_impressions).toLocaleString(); }
        catch(e){ bi = String(d.black_impressions); }
      }

      return `<tr>
        <td class="printer-cell">
          <a href="#" class="printer-link printer-name" data-ip="${d.ip||''}">${(d.printer_name||'').trim()||'(unnamed)'}</a>
          <div class="printer-meta">IP: ${d.ip||'N/A'}${d.location ? ` | LOC: ${String(d.location).split('/')[0].trim()}` : ''}</div>
        </td>
        <td class="nowrap"><span class="model-chip">${d.model||'Unknown'}</span></td>
        <td class="nowrap">${bi}</td>
        <td>${valueHTML(toner)}</td>
        <td>${valueHTML(drum)}</td>
        <td>${valueHTML(trR7)}</td>
        <td>${valueHTML(fuser)}</td>
        <td>${valueHTML(waste)}</td>
        <td>${valueHTML(beltcl)}</td>
        <td>${valueHTML(sbtr)}</td>
      </tr>`;
    }

    function tableHTML(rowsHTML, sortKey, sortDir){
      const thead = headerHTML(sortKey, sortDir);
      const tbody = `<tbody>${rowsHTML}</tbody>`;
      return `<div class="tableWrap"><table>${thead}${tbody}</table></div>`;
    }

    function pctBar(p){
      if(!(typeof p==='number') || p<0) return '';
      const w = Math.max(0, Math.min(100, p));
      return `<div class="muted" style="margin-bottom:4px">${p.toFixed(1)}%</div>
              <div class="bar"><div style="width:${w}%;background:${pctColor(p)}"></div></div>`;
    }
    function cmykRow(title, m){
      const order = ["K","C","M","Y"].filter(k => Object.prototype.hasOwnProperty.call(m,k));
      if(!order.length) return '';
      const nameMap = (title === 'Toner') ? { K: 'Black' } : {};
      const cells = order.map(k=>{
        const tag = nameMap[k] || k;
        return `<div class="metric"><div class="label">${title} ${tag}</div>${pctBar(m[k])}</div>`;
      }).join('');
      return `<div class="kv">${cells}</div>`;
    }
    function c415CardHTML(d){
      if(!d || String(d.model).toUpperCase()!=='C415') return '';
      const c = d.c415 || {};
      const t = c.toner || {};
      const drum = c.drum || {};
      const waste = (typeof c.waste_toner==='number' && c.waste_toner>=0) ? pctBar(c.waste_toner) : '';
      const beltcl = (typeof c.belt_cleaner==='number' && c.belt_cleaner>=0) ? pctBar(c.belt_cleaner) : '';
      const belt = (typeof c.transfer_belt==='number' && c.transfer_belt>=0) ? pctBar(c.transfer_belt) : '';
      const tr = (typeof c.transfer_roller==='number' && c.transfer_roller>=0) ? pctBar(c.transfer_roller) : '';

      let bi = 'N/A';
      if(typeof d.black_impressions === 'number' && d.black_impressions >= 0){
        try{ bi = Number(d.black_impressions).toLocaleString(); } catch(e){ bi = String(d.black_impressions); }
      }

      return `
        <div class="c415-card">
          <div style="font-weight:800; font-size:16px">${d.printer_name || '(unnamed)'} - ${d.ip} <span class="muted">[C415]</span></div>
          <div class="muted" style="margin:4px 0 10px">Black Impressions: <strong>${bi}</strong></div>
          ${cmykRow('Toner', t)}
          ${cmykRow('Drum', drum)}
          <div class="kv">
            ${waste?`<div class="metric"><div class="label">Waste Toner Container</div>${waste}</div>`:''}
            ${beltcl?`<div class="metric"><div class="label">Transfer Belt Cleaner</div>${beltcl}</div>`:''}
            ${belt?`<div class="metric"><div class="label">Transfer Belt</div>${belt}</div>`:''}
            ${tr?`<div class="metric"><div class="label">Transfer Roller</div>${tr}</div>`:''}
          </div>
        </div>
      `;
    }

    function metricSortValue(item){
      if(!item) return -1;
      if(item.status){
        return String(item.status).toLowerCase()==='ok' ? 100 : -1;
      }
      const p = (typeof item.percent==='number' && item.percent>=0) ? item.percent : -1;
      return p;
    }

    function buildRow(d){
      const items = d.items || [];
      const model = (d.model||'').toUpperCase();
      const toner   = pickBest(items,'toner');
      const drum    = pickBest(items,'drum');
      const tr      = pickBest(items,'transfer_roller');
      const fuser   = pickBest(items,'fuser');
      const waste   = pickBest(items,'waste_toner');
      const beltcl  = pickBest(items,'belt_cleaner');

      let trR7 = null, sbtr = null;
      if(tr){ if(model==='B8155') sbtr = tr; else trR7 = tr; }

      const bi = (typeof d.black_impressions === 'number' && d.black_impressions >= 0) ? d.black_impressions : -1;

      return {
        raw: d,
        sort: {
          ip: (d.ip||'').toLowerCase(),
          printer: (d.printer_name||'').toLowerCase(),
          model: (d.model||'').toLowerCase(),
          black: bi,
          toner: metricSortValue(toner),
          drum: metricSortValue(drum),
          tr_r7: metricSortValue(trR7),
          fuser: metricSortValue(fuser),
          waste: metricSortValue(waste),
          beltcl: metricSortValue(beltcl),
          sbtr: metricSortValue(sbtr),
        }
      };
    }

    function rowIsBelowThreshold(d){
      const items = d.items || [];
      const KEYS = ['toner','drum','transfer_roller','fuser','waste_toner','belt_cleaner'];
      const T = Number(state.threshold) || 10;
      for(const k of KEYS){
        const it = pickBest(items, k);
        if(it && typeof it.percent==='number' && it.percent>=0 && it.percent <= T){
          return true;
        }
      }
      return false;
    }

    function filteredRows(){
      const f = state.filter.trim().toLowerCase();
      let rows = state.rows;
      if(f){
        rows = rows.filter(r=> r.sort.ip.includes(f) || r.sort.printer.includes(f) || r.sort.model.includes(f));
      }
      if(state.onlyLow){
        rows = rows.filter(r => rowIsBelowThreshold(r.raw));
      }
      return rows;
    }

    function sortedRows(rows){
      const k = state.sortKey;
      const dir = state.sortDir === 'asc' ? 1 : -1;
      return rows.slice().sort((a,b)=>{
        const va = a.sort[k], vb = b.sort[k];
        if(typeof va === 'string' || typeof vb === 'string'){
          return String(va).localeCompare(String(vb)) * dir;
        }
        return ((va||0) - (vb||0)) * dir;
      });
    }

    function renderTable(){
      const out=document.getElementById('out');
      const rows1 = filteredRows();
      const rows2 = sortedRows(rows1);

      if(!rows2.length){
        out.className='empty';
        out.innerHTML = state.rows.length ? 'No printers match the current filters.' : 'No data yet. Run a query to see results.';
        return;
      }

      const htmlRows = rows2.map(r=>pivotRowHTML(r.raw)).join('');
      out.className='';
      out.innerHTML = tableHTML(htmlRows, state.sortKey, state.sortDir);

      const ths = out.querySelectorAll('thead th');
      ths.forEach(th=>{
        th.addEventListener('click', ()=>{
          const key = th.getAttribute('data-key');
          if(!key) return;
          if(state.sortKey===key){
            state.sortDir = (state.sortDir==='asc')?'desc':'asc';
          }else{
            state.sortKey = key;
            state.sortDir = (key==='ip' || key==='printer' || key==='model') ? 'asc' : 'desc';
          }
          renderTable();
        });
      });
    }

    document.getElementById('filter').addEventListener('input', (e)=>{
      state.filter = e.target.value || '';
      renderTable();
    });

    function setActiveSection(id){
      document.querySelectorAll('[data-section-link]').forEach(a => {
        a.classList.toggle('active', a.getAttribute('data-section-link') === id);
      });
    }

    document.querySelectorAll('[data-section-link]').forEach(a => {
      a.addEventListener('click', () => setActiveSection(a.getAttribute('data-section-link')));
    });

    const backTop = document.getElementById('backTop');
    if(backTop){
      backTop.addEventListener('click', () => window.scrollTo({ top: 0, behavior: 'smooth' }));
      window.addEventListener('scroll', () => {
        backTop.classList.toggle('show', window.scrollY > 360);
        const ids = ['results', 'summary', 'lookup'];
        const current = ids.find(id => {
          const el = document.getElementById(id);
          return el && el.getBoundingClientRect().top < 140;
        }) || 'lookup';
        setActiveSection(current);
      }, { passive: true });
    }

    const btnClear = document.getElementById('btnClear');
    if(btnClear){
      btnClear.addEventListener('click', () => {
        state.rows = [];
        state.filter = '';
        state.onlyLow = false;
        document.getElementById('filter').value = '';
        document.getElementById('onlyLow').checked = false;
        document.getElementById('extras').innerHTML = '';
        document.querySelectorAll('.query-status-msg').forEach(el => el.remove());
        updateMetrics([]);
        renderTable();
        document.getElementById('lookup').scrollIntoView({ behavior: 'smooth', block: 'start' });
      });
    }

    async function q(url){
  try {
    const r = await fetch(url);
    return await r.json();
  } catch(e) {
    console.error(e);
    return {error:'Network error'};
  }
}


    function buildQS(extra){ const p = new URLSearchParams({...extra}); return p.toString(); }

document.getElementById('f').addEventListener('submit', async (e)=>{
  e.preventDefault();

  // Clear previous completion messages.
  document.querySelectorAll('.query-status-msg').forEach(el => el.remove());

  // === Crear barra de progreso con spinner + contador ===
  const wrap = document.createElement('div');
  wrap.className = 'query-status-msg'; // <-- identificador para limpiarlo luego
  wrap.style.cssText = 'margin-top:10px;text-align:center;font-weight:700;color:#0f172a;font-size:13px;display:flex;flex-direction:column;align-items:center;gap:6px;';

  const topRow = document.createElement('div');
  topRow.style.cssText = 'display:flex;align-items:center;gap:8px;';

  const spinner = document.createElement('div');
  spinner.style.cssText = 'width:16px;height:16px;border:3px solid #cbd5e1;border-top-color:#2563eb;border-radius:50%;animation:spin .8s linear infinite;';
  const keyframes = document.createElement('style');
  keyframes.textContent = '@keyframes spin{to{transform:rotate(360deg)}}';
  if (!document.head.querySelector('style[data-spin]')) {
    keyframes.setAttribute('data-spin','1');
    document.head.appendChild(keyframes);
  }

  const label = document.createElement('div');
  label.textContent = 'Starting query...';
  topRow.appendChild(spinner);
  topRow.appendChild(label);

  const bar = document.createElement('div');
  bar.style.cssText = 'width:100%;background:#e2e8f0;border-radius:10px;overflow:hidden;box-shadow:inset 0 1px 3px rgba(0,0,0,.1)';
  const inner = document.createElement('div');
  inner.style.cssText = 'height:10px;width:0;background:#2563eb;transition:width .15s ease';
  bar.appendChild(inner);

  wrap.appendChild(topRow);
  wrap.appendChild(bar);
  document.querySelector('.sticky-card').appendChild(wrap);

  // === Determinar impresoras totales ===
  let total = 1;
  let ips = [];
  if (listmode.checked) {
    const ipsTxt = document.getElementById('ips').value || '';
    ips = ipsTxt.split(/[,\n\r\t\s]+/).map(s=>s.trim()).filter(Boolean);
    total = ips.length;
    if (!total) { showToast('Enter at least one IP.','error'); wrap.remove(); return; }
  } else {
    const ip = document.getElementById('ip').value.trim();
    if (!ip) { showToast('Enter an IP.','error'); wrap.remove(); return; }
    ips = [ip];
  }

  let pct = 0;
  const anim = setInterval(()=>{
    if (pct < 95) pct = Math.min(95, pct + Math.random() * 5);
    inner.style.width = pct + '%';
    label.textContent = `Querying ${ips.length > 1 ? ips.length + ' printers' : 'printer'}... ${Math.round(pct)}%`;
  }, 400);

  try {
    let data;
    if (listmode.checked) {
      const qs = buildQS({ ips: ips.join(',') });
      data = await q(`/api/supplies_list?${qs}`);
    } else {
      const qs = buildQS({ ip: ips[0] });
      data = await q(`/api/supplies?${qs}`);
    }

    // === Finish ===
    clearInterval(anim);
    spinner.remove();
    inner.style.width = '100%';
    label.textContent = `Completed ${ips.length} ${ips.length===1?'printer':'printers'} successfully`;

    bar.remove();
    wrap.style.marginTop = '12px';
    wrap.style.fontSize = '14px';
    wrap.style.color = '#166534'; // verde

    if (data.error) { showToast(data.error, 'error'); return; }
    prepare(data);

  } catch(err) {
    clearInterval(anim);
    wrap.remove();
    showToast('Query failed','error');
  }
});



    function prepare(data){
      const arr = Array.isArray(data.results) ? data.results : [data];
      const main = arr.filter(d => String(d.model||'').toUpperCase()!=='C415');
      state.rows = main.map(buildRow);
      updateMetrics(arr);
      renderTable();
      const cards = arr.filter(d=>String(d.model||'').toUpperCase()==='C415').map(c415CardHTML).join('');
      document.getElementById('extras').innerHTML = cards ? `<div class="section-title" style="margin-top:18px">C415 Details</div>${cards}` : '';
    }// ====== Printer Info Modal (Address + Business from sysLocation) ======
const pm = document.getElementById('printerModal');
const pmBackdrop = document.getElementById('pmBackdrop');
const pmCloseX = document.getElementById('pmCloseX');
const pmCloseBtn = document.getElementById('pmCloseBtn');
const pmSubtitle = document.getElementById('pmSubtitle');
const pmAddress = document.getElementById('pmAddress');
const pmBusiness = document.getElementById('pmBusiness');

function splitLocation(loc){
  const raw = String(loc || '').trim();
  if(!raw) return { address: 'N/A', business: 'N/A' };
  // split by first "/" only
  const parts = raw.split('/');
  if(parts.length >= 2){
    const address = parts[0].trim() || 'N/A';
    const business = parts.slice(1).join('/').trim() || 'N/A';
    return { address, business };
  }
  return { address: raw, business: 'N/A' };
}

function openPrinterModalByIP(ip){
  const row = (state.rows || []).find(r => (r.raw && String(r.raw.ip) === String(ip)));
  if(!row || !row.raw){
    showToast('Printer not found in current results.', 'warn');
    return;
  }
  const d = row.raw;
  const loc = splitLocation(d.location || '');
  pmSubtitle.textContent = `${(d.printer_name||'(unnamed)').trim()} - ${d.ip || ''} ${d.model ? ('['+d.model+']') : ''}`;
  pmAddress.textContent = loc.address;
  pmBusiness.textContent = loc.business;
  pm.style.display = 'block';
  document.body.style.overflow = 'hidden';
}

function closePrinterModal(){
  pm.style.display = 'none';
  document.body.style.overflow = '';
}

// Event delegation for printer links (works after re-render)
document.addEventListener('click', (ev)=>{
  const a = ev.target && ev.target.closest ? ev.target.closest('a.printer-link') : null;
  if(!a) return;
  ev.preventDefault();
  const ip = a.getAttribute('data-ip') || '';
  if(!ip){ showToast('Missing IP for this row.', 'error'); return; }
  openPrinterModalByIP(ip);
});

pmBackdrop.addEventListener('click', closePrinterModal);
pmCloseX.addEventListener('click', closePrinterModal);
pmCloseBtn.addEventListener('click', closePrinterModal);
document.addEventListener('keydown', (e)=>{ if(e.key === 'Escape' && pm.style.display === 'block') closePrinterModal(); });



    // Export con verificaciÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â³n y toast
    document.getElementById('btnXLSX').addEventListener('click', ()=>{
      if (!state.rows || state.rows.length === 0) { showToast('Run a query first.','warn'); return; }
      const visible = filteredRows();
      if (!visible || visible.length === 0) { showToast('No data to export','warn'); return; }

      const pct_max = state.onlyLow ? state.threshold : '';
      const text = state.filter || '';

      if (listmode.checked) {
        const ipsTxt = document.getElementById('ips').value||'';
        const ips = ipsTxt.split(/[,\n\r\t\s]+/).map(s=>s.trim()).filter(Boolean);
        if(!ips.length){ showToast('Enter at least one IP.','error'); return; }
        const qs = new URLSearchParams({ips: ips.join(','), pct_max, text}).toString();
        window.location = `/api/export_xlsx_list_pivot?${qs}`;
      } else {
        const ip = (document.getElementById('ip').value||'').trim();
        if(!ip){ showToast('Enter an IP.','error'); return; }
        const qs = new URLSearchParams({ip, pct_max, text}).toString();
        window.location = `/api/export_xlsx_pivot?${qs}`;
      }
    });
  </script>
</body>
</html>
'''
    return html.replace("__USER_MENU__", user_menu).replace("__LANG__", _lang())


# ---------------------- JSON API (web) --------------------
def _parse_ips_param() -> List[str]:
    raw = request.args.get("ips", "") or ""
    ips = [p.strip() for p in raw.replace("\r", "\n").replace(",", "\n").split("\n")]
    return [p for p in ips if p]

def _get_timeout_param() -> int:
    # UI no expone timeout; backend mantiene default
    return _parse_int(request.args.get("timeout"), TIMEOUT_DEFAULT)

@app.get("/api/supplies")
@login_required("api_supplies")
def api_supplies():
    ip = request.args.get("ip")
    if not ip: return jsonify({"error": "Missing 'ip' parameter"}), 400
    community = request.args.get("community", DEFAULT_COMMUNITY)
    timeout = _get_timeout_param()
    try:
        data = get_cached(ip, community, timeout, TTL_DEFAULT)
        return jsonify(data)
    except Exception as e:
        return jsonify({"ip": ip, "error": str(e)}), 500

@app.get("/api/supplies_list")
@login_required("api_supplies_list")
def api_supplies_list():
    """
    Consulta mÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Âºltiple optimizada:
    - B7135 ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¾ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ modo concurrente limitado (mÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¡x 4 ChromeDriver)
    - Otros modelos ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¾ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ threading normal (rÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¡pido)
    """
    import threading
    from queue import Queue

    ips = _parse_ips_param()
    if not ips:
        return jsonify({"error": "Missing 'ips' parameter with at least one IP"}), 400

    community = request.args.get("community", DEFAULT_COMMUNITY)
    timeout = _get_timeout_param()
    results = []

    b7135_ips, other_ips = [], []

    # Detectar modelos primero (solo SNMP, muy rÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¡pido)
    print("[INIT] Detectando modelos...")
    for ip in ips:
        try:
            model = _snmp_get_model(ip, community, timeout)
            if model == "B7135":
                b7135_ips.append(ip)
            else:
                other_ips.append(ip)
        except Exception as e:
            print(f"[MODEL DETECT ERROR] {ip}: {e}")
            other_ips.append(ip)

    # --- Consulta paralela para los no-B7135 ---
    def worker_generic(ip):
        try:
            data = get_cached(ip, community, timeout, TTL_DEFAULT)
            results.append(data)
        except Exception as e:
            print(f"[GENERIC THREAD ERROR] {ip}: {e}")
            results.append({"ip": ip, "error": str(e)})

    threads = []
    for ip in other_ips:
        t = threading.Thread(target=worker_generic, args=(ip,), daemon=True)
        threads.append(t)
        t.start()

    # --- Cola controlada para B7135 (mÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¡x 4 en paralelo) ---
    q = Queue()
    for ip in b7135_ips:
        q.put(ip)

    def worker_b7135():
        while not q.empty():
            ip = q.get()
            try:
                print(f"[B7135] Consultando {ip} ...")
                data = get_cached(ip, community, timeout, TTL_DEFAULT)
                results.append(data)
            except Exception as e:
                print(f"[B7135 ERROR] {ip}: {e}")
                results.append({"ip": ip, "error": str(e)})
            finally:
                q.task_done()
                time.sleep(1.5)  # descanso corto entre Seleniums

    # Lanzar 4 hilos Selenium simultÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¡neos
    max_parallel = 4
    for _ in range(min(max_parallel, len(b7135_ips))):
        t = threading.Thread(target=worker_b7135, daemon=True)
        t.start()
        threads.append(t)

    # Esperar todos
    for t in threads:
        t.join()

    print(f"[DONE] Total procesadas: {len(results)} impresoras")
    return jsonify({"results": results})





# ---------------------- Export XLSX -----------------
CSV_HEADERS = [
    "IP","Printer","Model","Black Impressions",
    "Toner","Drum/Imaging Unit",
    "Transfer Roller R7","Fuser R8",
    "Waste Toner Container","Transfer Belt Cleaner",
    "Second Bias Transfer Roll",
    "Exported At"
]

def _pick_best(items: List[Dict[str, Any]], category: str) -> Dict[str, Any]:
    cands = [it for it in items if it.get("category") == category]
    if not cands: return {}
    return max(cands, key=lambda it: (it.get("percent") if isinstance(it.get("percent"), (int, float)) else -1))

def pivot_row_for_xlsx(data: Dict[str, Any], exported_at: str) -> List[Any] | None:
    model = (data.get("model") or "").upper()
    if model == "C415":
        return None
    items = data.get("items", [])

    tr_item    = _pick_best(items, "transfer_roller")
    fuser_item = _pick_best(items, "fuser")
    toner_item = _pick_best(items, "toner")
    drum_item  = _pick_best(items, "drum")
    waste_item = _pick_best(items, "waste_toner")
    beltcl_it  = _pick_best(items, "belt_cleaner")

    def pct_num(it):
        if not it: return None
        if "status" in it and it["status"]:
            return it["status"]
        p = it.get("percent")
        if isinstance(p, (int, float)) and p >= 0:
            return float(p) / 100.0
        return None

    bi = data.get("black_impressions")
    bi_val = bi if isinstance(bi, int) and bi >= 0 else None

    tr_r7 = None
    sbtr  = None
    if tr_item:
        if model == "B8155": sbtr = pct_num(tr_item)
        else: tr_r7 = pct_num(tr_item)

    return [
        data.get("ip",""),
        data.get("printer_name",""),
        data.get("model",""),
        bi_val,
        pct_num(toner_item),
        pct_num(drum_item),
        tr_r7,
        pct_num(fuser_item),
        pct_num(waste_item),
        pct_num(beltcl_it),
        sbtr,
        exported_at
    ]

def _xlsx_apply_styles(ws):
    header_fill = PatternFill("solid", fgColor="F3F3F3")
    bold = Font(bold=True)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, title in enumerate(CSV_HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.fill = header_fill; c.font = bold; c.alignment = Alignment(vertical="center"); c.border = border
    widths = [14, 22, 24, 18, 12, 16, 20, 14, 22, 22, 26, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _xlsx_apply_grid(ws, start_row, start_col, end_row, end_col):
    thin = Side(style="thin", color="DDDDDD")
    medium = Side(style="medium", color="AAAAAA")
    for r in range(start_row, end_row+1):
        for c in range(start_col, end_col+1):
            left   = medium if c == start_col else thin
            right  = medium if c == end_col   else thin
            top    = medium if r == start_row else thin
            bottom = medium if r == end_row   else thin
            cell = ws.cell(row=r, column=c)
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

def _xlsx_format_body(ws, nrows):
    # EVITA aplicar formatos cuando no hay filas de datos
    if nrows < 2:
        return
    for r in range(2, nrows+1):
        ws.cell(row=r, column=4).number_format = '#,##0'
    for r in range(2, nrows+1):
        ws.cell(row=r, column=4).number_format = '#,##0'
    for c in range(5, 12):
        for r in range(2, nrows+1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, float):
                cell.number_format = '0%'

    green = Color(rgb="FF63BE7B")
    for c in range(5, 12):
        col_letter = get_column_letter(c)
        rng = f"{col_letter}2:{col_letter}{nrows}"
        ws.conditional_formatting.add(rng, DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=green, showValue=True))

    ok_fill  = PatternFill("solid", fgColor="E7F7EE")
    bad_fill = PatternFill("solid", fgColor="FDECEA")
    ws.conditional_formatting.add(f"G2:G{nrows}", CellIsRule(operator='equal', formula=['\"OK\"'], stopIfTrue=True, fill=ok_fill))
    ws.conditional_formatting.add(f"G2:G{nrows}", CellIsRule(operator='equal', formula=['\"Past end of life\"'], stopIfTrue=True, fill=bad_fill))
    ws.conditional_formatting.add(f"H2:H{nrows}", CellIsRule(operator='equal', formula=['\"OK\"'], stopIfTrue=True, fill=ok_fill))
    ws.conditional_formatting.add(f"H2:H{nrows}", CellIsRule(operator='equal', formula=['\"Past end of life\"'], stopIfTrue=True, fill=bad_fill))

    # sombreado warn/alert (independiente del UI)
    warn_fill  = PatternFill("solid", fgColor="FFF7ED")
    alert_fill = PatternFill("solid", fgColor="FEF2F2")
    for c in range(5, 12):
        col_letter = get_column_letter(c)
        rng = f"{col_letter}2:{col_letter}{nrows}"
        ws.conditional_formatting.add(rng, CellIsRule(operator='lessThanOrEqual', formula=[str(XLSX_ALERT_FRAC)], stopIfTrue=False, fill=alert_fill))
        ws.conditional_formatting.add(rng, CellIsRule(operator='lessThanOrEqual', formula=[str(XLSX_WARN_FRAC)],  stopIfTrue=False, fill=warn_fill))

    _xlsx_apply_grid(ws, 1, 1, nrows, len(CSV_HEADERS))

# ---------- Extra sheet: C415 Details ----------
def _xlsx_apply_styles_c415(ws, headers):
    header_fill = PatternFill("solid", fgColor="F3F3F3")
    bold = Font(bold=True)
    for col, title in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.fill = header_fill; c.font = bold; c.alignment = Alignment(vertical="center")
    widths = [14, 22, 10, 18] + [12]*12 + [20]
    for i, w in enumerate(widths[:len(headers)], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _add_c415_sheet(wb: Workbook, datas: List[Dict[str, Any]], exported_at: str):
    rows = [d for d in datas if (d.get("model") or "").upper() == "C415"]
    if not rows: return
    ws = wb.create_sheet(title="C415 Details")
    HEAD = [
        "IP","Printer","Model","Black Impressions",
        "Toner Black","Toner C","Toner M","Toner Y",
        "Drum K","Drum C","Drum M","Drum Y",
        "Waste Toner","Transfer Belt Cleaner","Transfer Belt","Transfer Roller",
        "Exported At"
    ]
    _xlsx_apply_styles_c415(ws, HEAD)

    rcount = 1
    def frac(v):
        try: v = float(v); return None if v < 0 else v/100.0
        except Exception: return None

    for d in rows:
        c = d.get("c415", {}) or {}
        t = c.get("toner", {}) or {}
        dr = c.get("drum", {}) or {}
        bi = d.get("black_impressions")
        bi_val = bi if isinstance(bi, int) and bi >= 0 else None

        row = [
            d.get("ip",""), d.get("printer_name",""), d.get("model",""), bi_val,
            frac(t.get("K", -1)), frac(t.get("C", -1)), frac(t.get("M", -1)), frac(t.get("Y", -1)),
            frac(dr.get("K", -1)), frac(dr.get("C", -1)), frac(dr.get("M", -1)), frac(dr.get("Y", -1)),
            frac(c.get("waste_toner", -1)),
            frac(c.get("belt_cleaner", -1)),
            frac(c.get("transfer_belt", -1)),
            frac(c.get("transfer_roller", -1)),
            exported_at
        ]
        ws.append(row); rcount += 1

    for r in range(2, rcount+1):
        ws.cell(row=r, column=4).number_format = '#,##0'
        for c in range(5, 17):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, float):
                cell.number_format = '0%'

    green = Color(rgb="FF63BE7B")
    for c in range(5, 17):
        col_letter = get_column_letter(c)
        rng = f"{col_letter}2:{col_letter}{rcount}"
        ws.conditional_formatting.add(rng, DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=green, showValue=True))

    warn_fill  = PatternFill("solid", fgColor="FFF7ED")
    alert_fill = PatternFill("solid", fgColor="FEF2F2")
    for c in range(5, 17):
        col_letter = get_column_letter(c)
        rng = f"{col_letter}2:{col_letter}{rcount}"
        ws.conditional_formatting.add(rng, CellIsRule(operator='lessThanOrEqual', formula=[str(XLSX_ALERT_FRAC)], stopIfTrue=False, fill=alert_fill))
        ws.conditional_formatting.add(rng, CellIsRule(operator='lessThanOrEqual', formula=[str(XLSX_WARN_FRAC)],  stopIfTrue=False, fill=warn_fill))

    _xlsx_apply_grid(ws, 1, 1, rcount, len(HEAD))

# --- Helpers para filtrar export ---
EXPORT_KEYS = ['toner','drum','transfer_roller','fuser','waste_toner','belt_cleaner']

def _any_metric_leq_pct(items: List[Dict[str, Any]], pct_max: float) -> bool:
    for k in EXPORT_KEYS:
        cand = _pick_best(items, k)
        if not cand:
            continue
        p = cand.get("percent")
        if isinstance(p, (int, float)) and p >= 0 and p <= pct_max:
            return True
        if k in ('fuser','transfer_roller') and isinstance(cand.get("status"), str):
            if cand["status"].lower() != "ok":
                return True
    return False

def _row_matches_filters(data: Dict[str, Any], pct_max: Optional[float], text: str) -> bool:
    ip = (data.get("ip") or "").lower()
    printer = (data.get("printer_name") or "").lower()
    model = (data.get("model") or "").lower()
    if text:
        t = text.lower().strip()
        if t and not (t in ip or t in printer or t in model):
            return False
    if pct_max is not None:
        items = data.get("items") or []
        if not _any_metric_leq_pct(items, pct_max):
            return False
    return True

def _parse_pct_max_arg(arg_val: Optional[str]) -> Optional[float]:
    if arg_val is None:
        return None
    s = str(arg_val).strip()
    if not s:
        return None
    try:
        v = float(s)
        if v < 0: v = 0.0
        if v > 100: v = 100.0
        return v
    except Exception:
        return None

@app.get("/api/export_xlsx_pivot")
@login_required("export_xlsx_pivot")
def export_xlsx_pivot():
    ip = request.args.get("ip")
    if not ip: return jsonify({"error": "Missing 'ip' parameter"}), 400
    community = request.args.get("community", DEFAULT_COMMUNITY)
    timeout = _get_timeout_param()

    # filtros recibidos
    pct_max = _parse_pct_max_arg(request.args.get("pct_max"))
    text = (request.args.get("text") or "").strip()

    data = get_cached(ip, community, timeout, TTL_DEFAULT)

    exported_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    fname_stamp = datetime.now().strftime("%Y%m%d-%H%M")

    wb = Workbook()
    ws = wb.active; ws.title = "Supplies"
    _xlsx_apply_styles(ws)

    if _row_matches_filters(data, pct_max, text):
        row = pivot_row_for_xlsx(data, exported_at)
        if row: ws.append(row)

    _xlsx_format_body(ws, nrows=ws.max_row)



    # Hoja C415: aplica mismo filtro por consistencia
    if _row_matches_filters(data, pct_max, text):
        _add_c415_sheet(wb, [data], exported_at)

    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    fname = f"xerox_supplies_{ip.replace('.', '-')}_{fname_stamp}.xlsx"
    headers = {"Content-Type":"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
               "Content-Disposition":f'attachment; filename="{fname}"'}
    return Response(bio.getvalue(), headers=headers)

@app.get("/api/export_xlsx_list_pivot")
@login_required("export_xlsx_list_pivot")
def export_xlsx_list_pivot():
    ips = _parse_ips_param()
    if not ips: return jsonify({"error": "Missing 'ips' parameter with at least one IP"}), 400
    community = request.args.get("community", DEFAULT_COMMUNITY)
    timeout = _get_timeout_param()

    # filtros recibidos
    pct_max = _parse_pct_max_arg(request.args.get("pct_max"))
    text = (request.args.get("text") or "").strip()

    exported_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    fname_stamp = datetime.now().strftime("%Y%m%d-%H%M")

    wb = Workbook()
    ws = wb.active; ws.title = "Supplies"
    _xlsx_apply_styles(ws)

    all_data = []
    for ip in ips:
        d = get_cached(ip, community, timeout, TTL_DEFAULT)
        if _row_matches_filters(d, pct_max, text):
            all_data.append(d)
            row = pivot_row_for_xlsx(d, exported_at)
            if row: ws.append(row)

    _xlsx_format_body(ws, nrows=ws.max_row)


    _add_c415_sheet(wb, all_data, exported_at)

    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    fname = f"xerox_supplies_list_{fname_stamp}.xlsx"
    headers = {"Content-Type":"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
               "Content-Disposition":f'attachment; filename="{fname}"'}
    return Response(bio.getvalue(), headers=headers)

# --------------------- Diagnostics ------------------------
@app.get("/api/dump_supplies")
@login_required("api_dump_supplies")
def api_dump_supplies():
    ip = request.args.get("ip")
    if not ip: return jsonify({"error": "Missing 'ip' parameter"}), 400
    community = request.args.get("community", DEFAULT_COMMUNITY)
    timeout = _get_timeout_param()
    try:
        model = _snmp_get_model(ip, community, timeout)
        descs  = _snmp_column_map(ip, community, DESC_BASE,  timeout=timeout)
        maxs   = _snmp_column_map(ip, community, MAX_BASE,   timeout=timeout)
        levels = _snmp_column_map(ip, community, LEVEL_BASE, timeout=timeout)

        rows = []
        for idx in sorted(set(descs) | set(maxs) | set(levels), key=lambda x: int(x)):
            desc = descs.get(idx) or ""
            pct  = _safe_pct(levels.get(idx), maxs.get(idx))
            rows.append({
                "index": int(idx),
                "desc": desc,
                "category": _categorize(desc),
                "level": levels.get(idx),
                "max": maxs.get(idx),
                "percent": pct
            })
        return jsonify({"ip": ip, "model": model, "rows": rows})
    except Exception as e:
        return jsonify({"ip": ip, "error": str(e)}), 500

@app.get("/api/debug")
@login_required("api_debug")
def api_debug():
    ip = request.args.get("ip")
    if not ip: return jsonify({"error": "Missing 'ip' parameter"}), 400
    community = request.args.get("community", DEFAULT_COMMUNITY)
    timeout = _get_timeout_param()
    try:
        sys_descr = _snmp_get(ip, community, SYS_DESCR, timeout)
        prt_name  = _snmp_get(ip, community, PRT_NAME, timeout)
        sys_name  = _snmp_get(ip, community, SYS_NAME, timeout)
        name = prt_name or sys_name
        model = _snmp_get_model(ip, community, timeout)
        descs  = _snmp_column_map(ip, community, DESC_BASE,  timeout=timeout)
        maxs   = _snmp_column_map(ip, community, MAX_BASE,   timeout=timeout)
        levels = _snmp_column_map(ip, community, LEVEL_BASE, timeout=timeout)
        sample = []
        for idx in sorted(set(list(descs)[:5] + list(maxs)[:5] + list(levels)[:5]), key=lambda x: int(x)):
            sample.append({"index": int(idx), "desc": descs.get(idx), "max": maxs.get(idx), "level": levels.get(idx)})
        black_impr = _get_black_impressions(ip, community, timeout)
        return jsonify({
            "ip": ip, "community_used": community, "model": model,
            "printer_name": name, "sys_name": sys_name, "sys_descr": sys_descr,
            "counts": {"descs": len(descs), "maxs": len(maxs), "levels": len(levels)},
            "sample_rows": sample, "black_impressions": black_impr
        })
    except Exception as e:
        return jsonify({"ip": ip, "community_used": community, "error": str(e)}), 500

@app.get("/healthz")
def healthz():
    return jsonify({"ok": True})

# -------------------------- main --------------------------
if __name__ == "__main__":
    app.run(
        host=FLASK_HOST,
        port=FLASK_PORT,
        debug=False,
        use_reloader=False,
        threaded=True,
    )

